"""
solution.py 编写模板 — 适用于 Cell-Level / Sheet-Level Manipulation 任务

用法：
1. 从 task.md 确定指令类型和预期位置
2. 用 inspect_input.py（或 bash 内联）检查 input.xlsx 实际结构
3. 根据模式选择下方对应代码块，填入具体逻辑
4. 保存为 solution.py 并运行 python run_solution.py 验证
"""

# ============================================================
# 通用加载
# ============================================================
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from datetime import datetime

# INPUT_PATH / OUTPUT_PATH 由 run_solution.py 注入，不要自行定义
wb = openpyxl.load_workbook(INPUT_PATH)


# ============================================================
# 模式 A：Cell-Level — 时间格式化（如 J 列提取时间+AM/PM）
# ============================================================
def mode_time_formatting(ws):
    """从 datetime 列提取时间，格式化后写入目标列"""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        src = row[8]   # 列 I (0-based index) — 含 datetime
        tgt = row[9]   # 列 J — 写入格式化时间

        if isinstance(src.value, datetime):
            tgt.value = src.value.time().strftime('%I:%M:%S %p')  # 12h AM/PM
            tgt.number_format = 'HH:MM:SS AM/PM'
        # 若原为公式（如 '=RIGHT(I2,8)'），直接覆盖
        elif isinstance(src.value, str) and src.value.startswith('='):
            pass  # 不用处理，直接从上文 datetime 列提取


# ============================================================
# 模式 B：Sheet-Level — 跨 sheet 聚合（STA/RPA → COLLECTION）
# ============================================================
def mode_cross_sheet_aggregation(wb):
    """跨多 sheet 按 (TY, OR) 键聚合 SALE/RET 并计算 BALANCE"""
    from collections import defaultdict

    data_sheets = [s for s in wb.sheetnames if s != 'COLLECTION']
    agg = defaultdict(lambda: {'sale': 0, 'ret': 0})

    # 步骤 1：遍历所有数据 sheet，按 (TY, OR) 分组求和
    for sn in data_sheets:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            ty, or_ = row[2], row[3]
            sale = float(row[4]) if row[4] not in (None, '-', 0) else 0.0
            ret = float(row[5]) if row[5] not in (None, '-', 0) else 0.0
            agg[(ty, or_)]['sale'] += sale
            agg[(ty, or_)]['ret'] += ret

    # 步骤 2：写入 COLLECTION sheet
    coll_ws = wb['COLLECTION']
    existing_rows = coll_ws.max_row
    row_idx = existing_rows + 1

    # 从现有数据中查找 ITEM, BR（需根据实际结构调整）
    # ... 根据 task.md 描述编写具体匹配逻辑

    for (ty, or_), vals in agg.items():
        balance = vals['sale'] - vals['ret']

        # 写入 ... (列映射根据实际结构调整)
        coll_ws.cell(row=row_idx, column=3).value = ty
        coll_ws.cell(row=row_idx, column=4).value = or_
        coll_ws.cell(row=row_idx, column=5).value = vals['sale']
        coll_ws.cell(row=row_idx, column=6).value = vals['ret']

        bal_cell = coll_ws.cell(row=row_idx, column=7)
        if balance < 0:
            bal_cell.value = balance
            bal_cell.font = Font(color='FF0000')  # 红色负数
        elif balance == 0:
            bal_cell.value = '-'
        else:
            bal_cell.value = balance

        row_idx += 1

    # 步骤 3：批量格式化
    font_def = Font(name='Calibri', size=11, bold=False)
    align_left = Alignment(horizontal='left')
    align_right = Alignment(horizontal='right')

    # 表头
    for cell in coll_ws[1]:
        cell.font = font_def
        cell.alignment = align_left

    # 数据行按列设对齐
    for row in coll_ws.iter_rows(min_row=2, max_row=coll_ws.max_row):
        for c in [0, 4, 5, 6]:  # A/E/F/G 列
            row[c].alignment = align_right
        for c in [1, 2, 3]:     # B/C/D 列
            row[c].alignment = align_left


# ============================================================
# 保存
# ============================================================
# ws = wb.active
# mode_time_formatting(ws)
# mode_cross_sheet_aggregation(wb)

wb.save(OUTPUT_PATH)