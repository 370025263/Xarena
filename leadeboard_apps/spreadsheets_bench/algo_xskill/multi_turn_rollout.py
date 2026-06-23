#!/usr/bin/env python3
"""多轮纠错 rollout —— 针对单个 SpreadsheetBench 题目，驱动 claude CLI 多轮对话。

目的（why）
==========
当前 xskill 的训练轨迹是"单轮"的：solver 一次写出 solution.py 就结束，无论对错。
单轮轨迹里没有"我错了 → 我怎么定位 → 我怎么改对"的修正过程，蒸馏出来的技能
往往是净负面的（教坏 solver）。本脚本产出"错误 → 修正 → 成功"的多轮高质量轨迹：

  turn0  solver 读题 + 用 Skill 工具匹配技能 → 把解法写到 solution.py → 执行
  判错   用（复制自官方 evaluator 的）cell 对比逻辑判对错
  反馈   若错，由"人味反馈"（隐藏 golden 真实数值，只点出哪些 cell 错了 + 怎么想）引导
  续会话 claude 用 `--resume <session_id>` 续上**同一会话**修正
  循环   直到做对 或 到 max_turns

claude CLI 的多轮会话会被它自己写到 `$CLAUDE_CONFIG_DIR/projects/*.jsonl`，
xskill daemon 监听该目录即可把整段多轮对话入库、蒸馏成技能。

自包含约束（重要）
==================
- 只依赖 openpyxl + 标准库。**不 import** skillopt 的任何模块（怕拖重依赖如 torch）。
  cell 对比 / 反馈 / prompt 全部在本文件内本地实现（逻辑复制自下列参考）：
    * cell 对比口径   <- skillopt/envs/spreadsheetbench/evaluator.py
    * 隐藏 golden 反馈 <- skillopt/envs/spreadsheetbench/codegen_agent._build_eval_feedback
    * workspace / 路径 <- skillopt/envs/spreadsheetbench/rollout.py
    * claude CLI cmd 构造 + env 隔离 <- skillopt/model/codex_harness._run_claude_code_cli_exec
- fail-loud 但不崩循环：claude 没输出 / 没写 solution.py / session_id 解析失败 /
  执行报错——都记录到 result 并合理处理，不静默吞错、也不让整脚本 crash。

claude CLI `-p` 非交互 + `--resume` 多轮的不确定点（需容器实测）
===============================================================
见文件末 `RESUME_CAVEAT` 常量与 README 报告。简言之：`-p`(print/非交互)模式下
`--resume <sid> -- <prompt>` 是否把新 prompt 当作"同一会话的下一条 user 消息"
续上、并把整段对话追加进**同一个** projects/*.jsonl，需要在有 claude CLI 的容器里
实测验证。本脚本按"会续上同一会话、首轮 json 输出里能解析到 session_id"的假设实现。
"""
from __future__ import annotations

import argparse
import datetime
import json
import os
import subprocess
import sys
import traceback

import openpyxl


# ════════════════════════════════════════════════════════════════════════════
# 1) cell 对比逻辑 —— 逐字复制自 skillopt/envs/spreadsheetbench/evaluator.py
#    口径：数值 round 2 位；datetime.time 去微秒；datetime 转 Excel 序列号取整；
#    ""与None视为相等；类型不同即 FAIL。保证训练判对错与官方评测一致。
# ════════════════════════════════════════════════════════════════════════════

def _datetime_to_float(dt: datetime.datetime) -> float:
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def _transform_value(v):
    if isinstance(v, bool):
        # openpyxl 可能返回 Python bool；官方把 bool 当数值处理(round(float(True))==1.0)
        return round(float(v), 2)
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, datetime.datetime):
        return round(_datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def _compare_cell_value(v1, v2) -> bool:
    v1 = _transform_value(v1)
    v2 = _transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        return False
    return v1 == v2


def _col_num2name(n: int) -> str:
    name = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(65 + r) + name
    return name


def _col_name2num(name: str) -> int:
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def _parse_range(range_str: str):
    start_cell, end_cell = range_str.split(":")
    sc = "".join(ch for ch in start_cell if ch.isalpha())
    sr = "".join(ch for ch in start_cell if ch.isdigit())
    ec = "".join(ch for ch in end_cell if ch.isalpha())
    er = "".join(ch for ch in end_cell if ch.isdigit())
    return (_col_name2num(sc), int(sr)), (_col_name2num(ec), int(er))


def _generate_cell_names(range_str: str):
    if ":" not in range_str:
        return [range_str]
    (sc, sr), (ec, er) = _parse_range(range_str)
    cols = [_col_num2name(i) for i in range(sc, ec + 1)]
    return [f"{c}{r}" for c in cols for r in range(sr, er + 1)]


def _iter_answer_targets(answer_position: str, default_sheet: str):
    """把 answer_position 串拆成 [(sheet_name, [cell_name, ...]), ...]。

    answer_position 形如 "I12:I13" 或 "Sheet2!A1:B3" 或逗号分隔的多段；
    不含 '!' 时落到 default_sheet（gt 的第一个 sheet，与官方一致）。
    """
    targets = []
    for scr in (answer_position or "").split(","):
        scr = scr.strip()
        if not scr:
            continue
        if "!" in scr:
            sheet_name, cell_range = scr.split("!", 1)
            sheet_name = sheet_name.strip().strip("'\"")
        else:
            sheet_name = default_sheet
            cell_range = scr
        cell_range = cell_range.strip().strip("'\"")
        targets.append((sheet_name, _generate_cell_names(cell_range)))
    return targets


def evaluate_output(pred_path: str, gold_path: str, answer_position: str) -> dict:
    """对比 pred 与 gold 在 answer_position 处的 cell。

    返回 {"ok": bool, "reason": str, "wrong_cells": [{"cell","got"}...]}。
    wrong_cells 只含 pred 自己的值(got)，**不含 expected 真值**——供反馈生成时使用，
    天然避免 golden 泄漏。fail-loud：文件缺失/打不开都如实返回 ok=False + reason。
    """
    if not os.path.exists(pred_path):
        return {"ok": False, "reason": "output file does not exist", "wrong_cells": []}
    try:
        wb_gt = openpyxl.load_workbook(filename=gold_path, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=pred_path, data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "reason": f"load error: {e}", "wrong_cells": []}

    try:
        default_sheet = wb_gt.sheetnames[0]
        wrong_cells: list[dict] = []
        reason = ""
        for sheet_name, cell_names in _iter_answer_targets(answer_position, default_sheet):
            if sheet_name not in wb_proc.sheetnames:
                if not reason:
                    reason = f"worksheet not found in output: {sheet_name}"
                # 整个 sheet 缺失：把该 sheet 的目标 cell 全部记为错(got 缺失)
                for cn in cell_names:
                    wrong_cells.append({"cell": f"{sheet_name}!{cn}", "got": None})
                continue
            ws_gt = wb_gt[sheet_name]
            ws_proc = wb_proc[sheet_name]
            for cn in cell_names:
                cg = ws_gt[cn].value
                cp = ws_proc[cn].value
                if not _compare_cell_value(cg, cp):
                    wrong_cells.append({"cell": f"{sheet_name}!{cn}", "got": cp})
                    if not reason:
                        reason = f"value mismatch @ {sheet_name}!{cn}"
        ok = len(wrong_cells) == 0
        return {"ok": ok, "reason": "" if ok else reason, "wrong_cells": wrong_cells}
    finally:
        wb_gt.close()
        wb_proc.close()


# ════════════════════════════════════════════════════════════════════════════
# 2) workbook 预览 —— 复制自 codegen_agent._preview_workbook
# ════════════════════════════════════════════════════════════════════════════

def preview_workbook(path: str, max_rows: int = 5, max_cols: int = 20) -> str:
    """生成 workbook 前几行的文本预览（用于 turn0 的 input 预览）。"""
    wb = openpyxl.load_workbook(path, data_only=False)
    chunks: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            chunks.append(
                f"## Sheet: {sheet_name}  "
                f"(dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})"
            )
            for row in ws.iter_rows(
                min_row=1,
                max_row=min(ws.max_row, max_rows),
                max_col=min(ws.max_column, max_cols),
                values_only=False,
            ):
                cells = []
                for cell in row:
                    v = cell.value
                    if v is None:
                        cells.append(f"{cell.coordinate}=")
                    else:
                        s = str(v)
                        if len(s) > 40:
                            s = s[:37] + "..."
                        cells.append(f"{cell.coordinate}={s}")
                chunks.append(" | ".join(cells))
            if ws.max_row > max_rows:
                chunks.append(f"... ({ws.max_row - max_rows} more rows)")
            chunks.append("")
    finally:
        wb.close()
    return "\n".join(chunks)


# ════════════════════════════════════════════════════════════════════════════
# 3) 数据定位 —— 复制 rollout._find_test_cases 的命名约定（简化为单 case）
# ════════════════════════════════════════════════════════════════════════════

def load_dataset(data_root: str) -> list[dict]:
    """读 <data-root>/dataset.json（官方为 list[dict]）。"""
    path = os.path.join(data_root, "dataset.json")
    if not os.path.exists(path):
        raise FileNotFoundError(f"dataset.json not found at {path}")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        data = data.get("data") or list(data.values())
    return list(data)


def find_item(items: list[dict], item_id: str) -> dict:
    """按 str(id) 匹配（id 可能是 int 或 str）。找不到 fail-loud。"""
    want = str(item_id)
    for it in items:
        if str(it.get("id")) == want:
            return it
    raise KeyError(f"item id={item_id!r} not found in dataset.json")


def find_input_and_golden(task_dir: str) -> tuple[str, str]:
    """在 spreadsheet/<id>/ 找 input 和 golden。

    优先 *_init.xlsx + *_golden.xlsx（verified_400 的 1_<id>_init.xlsx 命名）；
    回退 initial.xlsx + golden.xlsx。fail-loud：找不到就抛错。
    """
    import glob as _glob
    inits = sorted(_glob.glob(os.path.join(task_dir, "*_init.xlsx")))
    for ip in inits:
        gp = ip.replace("_init.xlsx", "_golden.xlsx")
        if os.path.exists(gp):
            return ip, gp
    bare_init = os.path.join(task_dir, "initial.xlsx")
    bare_gold = os.path.join(task_dir, "golden.xlsx")
    if os.path.exists(bare_init) and os.path.exists(bare_gold):
        return bare_init, bare_gold
    raise FileNotFoundError(
        f"no (input, golden) pair found in {task_dir} "
        f"(looked for *_init.xlsx+*_golden.xlsx, initial.xlsx+golden.xlsx)"
    )


# ════════════════════════════════════════════════════════════════════════════
# 4) workspace 准备：run_solution.py 模板（注入 INPUT_PATH / OUTPUT_PATH）
# ════════════════════════════════════════════════════════════════════════════

def _build_run_solution_driver(input_path: str, output_path: str) -> str:
    """run_solution.py：定义 INPUT_PATH/OUTPUT_PATH，再 exec 同目录 solution.py。

    复制 codegen_agent._build_codex_driver 的思路：剥掉 solution.py 里用户自己写的
    INPUT_PATH/OUTPUT_PATH 赋值，强制用我们注入的路径，避免硬编码工作簿。
    """
    return (
        "import pathlib\n"
        "import re\n"
        "import sys\n"
        "import traceback\n\n"
        f"INPUT_PATH = {input_path!r}\n"
        f"OUTPUT_PATH = {output_path!r}\n"
        "code = pathlib.Path(__file__).with_name('solution.py').read_text(encoding='utf-8')\n"
        "# 剥掉用户在 solution.py 里自己写的 INPUT_PATH/OUTPUT_PATH 赋值\n"
        "code = re.sub(r'^\\s*(INPUT_PATH|OUTPUT_PATH)\\s*=\\s*.+$', '', code, flags=re.MULTILINE)\n"
        "g = {'__name__': '__main__', 'INPUT_PATH': INPUT_PATH, 'OUTPUT_PATH': OUTPUT_PATH}\n"
        "try:\n"
        "    exec(compile(code, 'solution.py', 'exec'), g, g)\n"
        "except Exception:\n"
        "    traceback.print_exc()\n"
        "    sys.exit(2)\n"
    )


def prepare_workspace(workspace: str, input_src: str) -> tuple[str, str]:
    """建 workspace：拷 input.xlsx、写 run_solution.py。返回 (input_path, output_path)。"""
    import shutil
    os.makedirs(workspace, exist_ok=True)
    input_path = os.path.join(workspace, "input.xlsx")
    output_path = os.path.join(workspace, "output.xlsx")
    shutil.copy2(input_src, input_path)
    driver = _build_run_solution_driver(input_path, output_path)
    with open(os.path.join(workspace, "run_solution.py"), "w", encoding="utf-8") as f:
        f.write(driver)
    return input_path, output_path


# ════════════════════════════════════════════════════════════════════════════
# 5) prompt 构造
# ════════════════════════════════════════════════════════════════════════════

def build_turn0_prompt(
    instruction: str,
    input_path: str,
    instruction_type: str,
    answer_position: str,
) -> str:
    """turn0：完整任务描述 + input 预览 + 用 Skill 工具匹配技能 + 写 solution.py 的指示。"""
    try:
        preview = preview_workbook(input_path)
    except Exception as e:  # noqa: BLE001
        preview = f"(failed to preview workbook: {e})"
    extra = ""
    if instruction_type:
        extra += f"\nInstruction type: {instruction_type}"
    if answer_position:
        extra += f"\nExpected answer position: {answer_position}"
    return (
        f"# Instruction\n{instruction}\n{extra}\n\n"
        f"# Input spreadsheet preview\n{preview}\n\n"
        "# Task\n"
        "- First, use the Skill tool to find and invoke any available skill whose "
        "description matches this spreadsheet task, then follow its guidance.\n"
        "- Inspect `input.xlsx` in this workspace if useful.\n"
        "- Write the final Python solution to `solution.py`. The script must read the "
        "workbook from the `INPUT_PATH` variable and write the modified workbook to "
        "`OUTPUT_PATH`. Preserve all other cells unchanged.\n"
        "- The preview may be truncated — do not hardcode row counts or assume the data "
        "ends at the last previewed row; iterate over all actual rows instead.\n"
        "- You may run `python run_solution.py` to validate locally.\n"
        "- In your final message, confirm `solution.py` was written and summarize the approach."
    )


def build_human_feedback(wrong_cells: list[dict], instruction: str) -> str:
    """第一阶段（--human-sim off）的"人味反馈"：golden-diff 直接生成。

    复制 codegen_agent._build_eval_feedback 的核心约束——**列出预测错的 cell + 你的值，
    但绝不写出 expected 真值**——但语气改成同事口吻而非机器报错。

    例（同事味）：
      "你 I12:I13 这几格的结果好像不太对。你算出来是 3 / 5 ……
       看看是不是 ……？另外检查一下 ……"
    而不是机器味的 "cell I12 got=3 expected=5"。
    """
    if not wrong_cells:
        # 不应发生（有反馈才进这里），但 fail-soft 给个通用提示
        return (
            "运行起来了，但结果还差点意思。再核对一下题目要求的那几个单元格，"
            "看看计算口径有没有偏差？"
        )

    # 把错的 cell 按 sheet!列 归类，方便用"这几格"的口语化指代
    cell_bits = []
    for wc in wrong_cells:
        got = wc.get("got")
        # got=None 用"是空的"表述，更像人话
        if got is None:
            cell_bits.append(f"{wc['cell']} 还是空的")
        else:
            cell_bits.append(f"{wc['cell']} 你填的是 {got!r}")

    listed = "；".join(cell_bits)
    # 给几个"同事会顺口提醒"的排查方向，引导但不给答案、不泄漏 expected
    hints = (
        "我猜可能的坑：(1) 日期/时间被当成文本处理了？(2) 公式没被求值——"
        "openpyxl 不会算 Excel 公式，得在 Python 里把结果算出来再写值；"
        "(3) 空格的填充/跳过逻辑、或者数值的四舍五入位数对不上；"
        "(4) 行数是不是只覆盖了预览里那几行、漏了后面的数据。"
    )
    return (
        f"嘿，跑通了但结果对不上。具体是这几格：{listed}。\n"
        f"对照下题目要求（“{instruction[:120]}{'…' if len(instruction) > 120 else ''}”），"
        f"{hints}\n"
        "你按这些方向再排查一下，把 `solution.py` 改对，然后我再帮你看。"
    )


def build_exec_error_feedback(err: str) -> str:
    """执行报错时的反馈（同样同事口吻，把报错贴给它让它改）。"""
    return (
        "诶，你这版 `solution.py` 跑的时候直接报错了：\n\n"
        f"```\n{err[:2500]}\n```\n\n"
        "先把这个错修掉吧，记得继续用 `INPUT_PATH` / `OUTPUT_PATH` 变量，别硬编码路径。"
    )


def build_no_solution_feedback() -> str:
    """没写出 solution.py 时的反馈。"""
    return (
        "我没在 workspace 里找到 `solution.py`。请把完整解法写进 `solution.py`，"
        "用 `INPUT_PATH` 读、写到 `OUTPUT_PATH`，可以先 `python run_solution.py` 自测。"
    )


# 第二阶段占位：用 LLM 把 golden-diff 翻译成更自然的人味反馈
def human_feedback_via_llm(wrong_cells, instruction, **kwargs):  # noqa: D401
    """[第二阶段实现] 用 LLM 生成更自然的人味反馈。

    设计：把 wrong_cells（只含 got，无 expected）+ instruction 喂给一个小模型，
    让它扮演"看了你输出但没看答案的同事"，产出引导性反馈。当前未实现，
    --human-sim on 时先回退到 golden-diff 版的 build_human_feedback，避免阻塞第一阶段。
    """
    raise NotImplementedError(
        "human_feedback_via_llm 是第二阶段特性，尚未实现；"
        "当前 --human-sim 仅作占位，运行时会回退到 build_human_feedback。"
    )


# ════════════════════════════════════════════════════════════════════════════
# 6) claude CLI 调用 —— 复制自 codex_harness._run_claude_code_cli_exec 的 cmd/env
# ════════════════════════════════════════════════════════════════════════════

# tools 必须含 Skill，否则 Claude Code 会把 skill_listing 整个剥掉 => 技能永远进不了预算。
_DEFAULT_TOOLS = "Read,Write,Edit,Bash,Skill"


def call_claude(
    *,
    claude_path: str,
    work_dir: str,
    chome: str,
    model: str,
    prompt: str,
    timeout: int,
    output_format: str,
    resume_session_id: str | None = None,
    tools: str = _DEFAULT_TOOLS,
) -> tuple[str, str, int]:
    """调一次 claude CLI。返回 (stdout, raw含stderr, returncode)。

    cmd 构造照 codex_harness._run_claude_code_cli_exec：
      claude -p --output-format <fmt> --permission-mode bypassPermissions
             --add-dir <work_dir> --tools <tools> --allowedTools <tools>
             --model <model> --setting-sources user,project
             [--resume <sid>] -- <prompt>
    env 继承 os.environ 并设 CLAUDE_CONFIG_DIR=<chome>（隔离到实验 config，
    实验技能在 <chome>/skills 下，--setting-sources user,project 才能加载到）。
    cwd=work_dir。

    fail-loud：超时/非零返回都如实带回 raw + returncode，由调用方判定，不在此吞错。
    """
    cmd = [
        claude_path,
        "-p",
        "--output-format", output_format,
        "--permission-mode", "bypassPermissions",
        "--add-dir", work_dir,
        "--tools", tools,
        "--allowedTools", tools,
        "--setting-sources", "user,project",
    ]
    if model:
        cmd += ["--model", model]
    if resume_session_id:
        # 续接同一会话（详见文件头与 RESUME_CAVEAT：-p 模式下的确切行为需容器实测）
        cmd += ["--resume", resume_session_id]
    cmd += ["--", prompt]

    run_env = dict(os.environ)
    run_env["CLAUDE_CONFIG_DIR"] = chome

    try:
        proc = subprocess.run(
            cmd,
            cwd=work_dir,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=run_env,
        )
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout or ""
        stderr = exc.stderr or ""
        raw = stdout
        if stderr:
            raw = f"{raw}\n[stderr]\n{stderr}" if raw else stderr
        return "", (raw or f"timeout after {timeout}s"), 124

    stdout = proc.stdout or ""
    stderr = proc.stderr or ""
    raw = stdout
    if stderr:
        raw = f"{raw}\n[stderr]\n{stderr}" if raw else stderr
    return stdout, raw, proc.returncode


def parse_session_id(stdout: str) -> str:
    """从 `--output-format json` 的 stdout 解析 session_id。

    claude CLI 的 -p json 输出是一个 JSON 对象（含 session_id/result 等字段）。
    fail-loud：解析不到返回空串，由调用方记录"session_id 解析失败"。
    """
    s = (stdout or "").strip()
    if not s:
        return ""
    # 先试整体 JSON
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            sid = obj.get("session_id") or obj.get("sessionId") or ""
            if sid:
                return str(sid)
    except (json.JSONDecodeError, ValueError):
        pass
    # 回退：逐行找可解析的 JSON 对象（stream-json / 多行场景）
    for line in s.splitlines():
        line = line.strip()
        if not line.startswith("{"):
            continue
        try:
            obj = json.loads(line)
        except (json.JSONDecodeError, ValueError):
            continue
        if isinstance(obj, dict):
            sid = obj.get("session_id") or obj.get("sessionId") or ""
            if sid:
                return str(sid)
    return ""


# ════════════════════════════════════════════════════════════════════════════
# 7) 主流程：多轮纠错循环
# ════════════════════════════════════════════════════════════════════════════

def run_multi_turn(args) -> dict:
    """对单个 item 跑多轮纠错 rollout，返回 result dict。"""
    result = {
        "item_id": str(args.item_id),
        "success": False,
        "turns_used": 0,
        "session_id": "",
        "per_turn": [],
        "fail_reason": "",
    }

    # ── setup：定位 item / 文件 / 准备 workspace ──────────────────────────
    items = load_dataset(args.data_root)
    item = find_item(items, args.item_id)
    instruction = item["instruction"]
    instruction_type = item.get("instruction_type", "")
    answer_position = item.get("answer_position", "")
    answer_sheet = item.get("answer_sheet", "")
    # answer_position 不带 sheet 但题目给了 answer_sheet 时，拼成 Sheet!Range（与官方一致）
    if answer_position and answer_sheet and "!" not in answer_position:
        answer_position_eval = f"{answer_sheet}!{answer_position}"
    else:
        answer_position_eval = answer_position
    result["answer_position"] = answer_position_eval

    sp = item.get("spreadsheet_path", f"spreadsheet/{args.item_id}")
    task_dir = sp if os.path.isabs(sp) else os.path.join(args.data_root, sp)
    input_src, golden_path = find_input_and_golden(task_dir)

    input_path, output_path = prepare_workspace(args.workspace, input_src)
    result["workspace"] = args.workspace

    session_id = ""

    # ── 多轮循环 ──────────────────────────────────────────────────────────
    for turn in range(args.max_turns):
        per = {"turn": turn, "exec_ok": False, "eval_ok": False,
               "n_wrong_cells": None, "claude_ok": False, "note": ""}

        # a) 组装本轮 prompt
        if turn == 0:
            prompt = build_turn0_prompt(
                instruction, input_path, instruction_type, answer_position_eval
            )
        else:
            prompt = feedback  # 上一轮末尾生成的人味反馈（见循环末）

        # b) 调 claude CLI。turn0 用 json 拿 session_id；turn>0 用 --resume 续会话。
        #    先删旧 output.xlsx，确保拿到的是本轮真实产物（不是上一轮残留）。
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except OSError:
                pass

        if turn == 0:
            stdout, raw, rc = call_claude(
                claude_path=args.claude_path, work_dir=args.workspace,
                chome=args.chome, model=args.model, prompt=prompt,
                timeout=args.timeout, output_format="json",
            )
            session_id = parse_session_id(stdout)
            result["session_id"] = session_id
            if not session_id:
                # session_id 解析失败：后续轮无法 --resume 续会话。
                # fail-loud 记录；本轮仍尝试执行（solution.py 可能已写出）。
                per["note"] = "session_id parse failed (cannot --resume in later turns)"
        else:
            if not session_id:
                # 没有 session_id，无法续会话——终止循环（fail-loud，不静默重开新会话）
                per["note"] = "no session_id; cannot resume — aborting multi-turn loop"
                result["per_turn"].append(per)
                if not result["fail_reason"]:
                    result["fail_reason"] = "no-session-id-to-resume"
                break
            stdout, raw, rc = call_claude(
                claude_path=args.claude_path, work_dir=args.workspace,
                chome=args.chome, model=args.model, prompt=prompt,
                timeout=args.timeout, output_format="text",
                resume_session_id=session_id,
            )

        per["returncode"] = rc
        per["claude_ok"] = bool((stdout or "").strip()) and rc == 0
        if not per["claude_ok"] and not per["note"]:
            per["note"] = f"claude produced no usable output (rc={rc})"

        # c) 检查 solution.py 是否写出
        solution_path = os.path.join(args.workspace, "solution.py")
        if not os.path.exists(solution_path):
            # 没写出 solution.py：本轮失败。还有后续轮则给反馈续上，否则终止。
            per["note"] = (per["note"] + "; " if per["note"] else "") + "no solution.py written"
            result["per_turn"].append(per)
            result["turns_used"] = turn + 1
            if turn + 1 >= args.max_turns:
                if not result["fail_reason"]:
                    result["fail_reason"] = "no-solution-py"
                break
            feedback = build_no_solution_feedback()
            continue

        # d) 执行 run_solution.py 产出 output.xlsx
        exec_ok, exec_err = _exec_run_solution(args.workspace, args.timeout)
        per["exec_ok"] = exec_ok

        if not exec_ok:
            # 执行报错：把报错反馈给它，进下一轮（或终止）
            result["per_turn"].append(per)
            result["turns_used"] = turn + 1
            if turn + 1 >= args.max_turns:
                if not result["fail_reason"]:
                    result["fail_reason"] = f"exec-error: {exec_err[:200]}"
                break
            feedback = build_exec_error_feedback(exec_err)
            continue

        # e) 判对错（复制来的 cell 对比逻辑）
        ev = evaluate_output(output_path, golden_path, answer_position_eval)
        per["eval_ok"] = ev["ok"]
        per["n_wrong_cells"] = len(ev["wrong_cells"])
        result["per_turn"].append(per)
        result["turns_used"] = turn + 1

        if ev["ok"]:
            # 做对了 —— 记 success、break
            result["success"] = True
            result["fail_reason"] = ""
            break

        # 做错了 —— 生成人味反馈进下一轮（若还有轮次）
        if turn + 1 >= args.max_turns:
            if not result["fail_reason"]:
                result["fail_reason"] = f"eval-mismatch: {ev['reason'][:200]}"
            break

        if args.human_sim:
            # 第二阶段：尝试 LLM 人味反馈；未实现则回退 golden-diff（注明）
            try:
                feedback = human_feedback_via_llm(ev["wrong_cells"], instruction)
            except NotImplementedError:
                feedback = build_human_feedback(ev["wrong_cells"], instruction)
        else:
            feedback = build_human_feedback(ev["wrong_cells"], instruction)

    return result


def _exec_run_solution(workspace: str, timeout: int) -> tuple[bool, str]:
    """执行 workspace/run_solution.py，产出 output.xlsx。返回 (ok, err)。

    用子进程跑（隔离 solver 代码的副作用/崩溃），fail-loud 把 stdout+stderr 带回。
    """
    run_solution = os.path.join(workspace, "run_solution.py")
    output_path = os.path.join(workspace, "output.xlsx")
    try:
        proc = subprocess.run(
            [sys.executable, run_solution],
            cwd=workspace,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return False, f"run_solution.py timeout after {timeout}s"
    if proc.returncode != 0:
        return False, (proc.stdout + "\n" + proc.stderr).strip()
    if not os.path.exists(output_path):
        return False, "run_solution.py finished but output.xlsx was not created"
    return True, ""


# claude CLI -p 模式 --resume 多轮的最大不确定点（需容器实测）
RESUME_CAVEAT = (
    "claude CLI 在 `-p`(print/非交互) 模式下 `--resume <sid> -- <prompt>` 的确切行为需实测：\n"
    " (1) 它是否把新 prompt 当作同一会话的下一条 user 消息续上、并把整段多轮对话\n"
    "     追加进同一个 $CLAUDE_CONFIG_DIR/projects/*.jsonl（而非新开一个 session 文件）？\n"
    " (2) `--output-format json` 首轮输出里 session_id 字段的确切键名（session_id / sessionId）；\n"
    " (3) --resume 是否要求与首轮相同的 --add-dir / cwd / model 才能成功续接；\n"
    " (4) bypassPermissions 下 Bash 自测(python run_solution.py)是否真能在该 workspace 跑通。\n"
    "本脚本按 (1) 成立、(2) 取 session_id 的假设实现；不成立时需改用 --continue 或落地会话目录。"
)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="单题 SpreadsheetBench 多轮纠错 rollout（驱动 claude CLI）"
    )
    parser.add_argument("--item-id", required=True, help="dataset.json 里的题目 id（int/str 皆可）")
    parser.add_argument("--data-root", default="/data", help="含 dataset.json 与 spreadsheet/ 的根目录")
    parser.add_argument("--chome", required=True, help="CLAUDE_CONFIG_DIR（隔离的 claude config 目录）")
    parser.add_argument("--claude-path", default="claude", help="claude CLI 可执行文件路径")
    parser.add_argument("--model", default="deepseek-v4-flash", help="模型名")
    parser.add_argument("--max-turns", type=int, default=5, help="最大轮数")
    parser.add_argument("--workspace", default=None, help="工作目录（默认 /tmp/mt_<item-id>）")
    parser.add_argument("--out", default=None, help="result json 落盘路径（可选）")
    parser.add_argument("--timeout", type=int, default=420, help="每轮 claude / 执行的超时秒数")
    parser.add_argument("--human-sim", action="store_true",
                        help="[第二阶段占位] 用 LLM 生成人味反馈；当前未实现，会回退 golden-diff")
    args = parser.parse_args()

    if not args.workspace:
        args.workspace = f"/tmp/mt_{args.item_id}"

    try:
        result = run_multi_turn(args)
    except Exception as e:  # noqa: BLE001
        # setup 阶段的硬错误（找不到 item / 文件等）也写成 result，整脚本不 crash
        result = {
            "item_id": str(args.item_id),
            "success": False,
            "turns_used": 0,
            "session_id": "",
            "per_turn": [],
            "fail_reason": f"setup-error: {type(e).__name__}: {e}",
            "error": traceback.format_exc(),
        }

    if args.out:
        os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

    # stdout 打印一行简洁结果
    print(
        f"[mt-rollout] id={result['item_id']} "
        f"success={result['success']} turns={result['turns_used']} "
        f"session={result.get('session_id') or '-'} "
        f"reason={result.get('fail_reason') or 'ok'}"
    )
    return 0 if result["success"] else 1


if __name__ == "__main__":
    sys.exit(main())
