import os
import uuid
import json
import re
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_jwt_extended import (
    JWTManager, create_access_token, get_jwt_identity, jwt_required, get_jwt
)
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import load_only
from werkzeug.utils import secure_filename

from kubernetes import client, config
from kubernetes.client.rest import ApiException

import click



# --- Flask 应用初始化 ---
app = Flask(__name__)

# 替换之前的 @app.after_request _add_cors
@app.after_request
def _add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, DELETE, OPTIONS"
    response.headers["Access-Control-Expose-Headers"] = "X-Filename, Content-Disposition"
    return response

# ✅ 新增：单独处理 OPTIONS 预检请求，否则 Flask 对未注册路由返回 404
@app.route("/api/<path:path>", methods=["OPTIONS"])
def _options_handler(path):
    return "", 204

# --- 核心配置 ---
app.config["JWT_SECRET_KEY"] = os.environ.get("JWT_SECRET_KEY", "your-super-secret-key-change-this")
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=int(os.environ.get("JWT_EXPIRATION_HOURS", 24)))
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///local_test.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
UPLOADS_HOST_PATH = os.environ.get("UPLOADS_HOST_PATH", "/models/leaderboard_uploads/uploads")
os.makedirs(UPLOADS_HOST_PATH, exist_ok=True)

# --- 扩展初始化 ---
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
jwt = JWTManager(app)

# --- Kubernetes API 配置 ---
k8s_core_v1 = None
k8s_batch_v1 = None
try:
    if os.environ.get("KUBERNETES_SERVICE_HOST"):
        config.load_incluster_config()
        app.logger.info("Loaded in-cluster K8s config.")
    else:
        config.load_kube_config()
        app.logger.info("Loaded local kubeconfig.")
    k8s_core_v1 = client.CoreV1Api()
    k8s_batch_v1 = client.BatchV1Api()
except config.ConfigException as e:
    app.logger.error(f"Could not load any K8s config: {e}")
    k8s_core_v1 = None
    k8s_batch_v1 = None

K8S_NAMESPACE = os.environ.get("K8S_NAMESPACE", "leaderboard")
K8S_SERVICE_ACCOUNT = os.environ.get("K8S_SERVICE_ACCOUNT", "leaderboard-api-sa")
K8S_JOB_TTL_SECONDS = int(os.environ.get("K8S_JOB_TTL_SECONDS", 3600))
K8S_JOB_ACTIVE_DEADLINE = int(os.environ.get("K8S_JOB_ACTIVE_DEADLINE", 60 * 60 * 12))

# ======================================================================
# 数据库模型
# ======================================================================


class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='participant')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)


class Leaderboard(db.Model):
    __tablename__ = 'leaderboards'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text, nullable=True)
    evaluator_image = db.Column(db.String(255), nullable=False)
    baseline_image = db.Column(db.String(255), nullable=True)
    resource_spec = db.Column(db.Text, nullable=False)  # JSON string
    version = db.Column(db.String(20), nullable=False)
    owner_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    owner = db.relationship('User', backref=db.backref('leaderboards', lazy=True))
    submissions = db.relationship('Submission', backref='leaderboard', lazy='dynamic', cascade="all, delete-orphan")

    # 新增：难度系数、SOTA 分数
    difficulty_factor = db.Column(db.Float, nullable=False, default=1.0)
    sota_score = db.Column(db.Float, nullable=True)  # 可空（未设 SOTA 时不计分）

    # 新增：每个榜单要求 algo 必须提供的 ENV（JSON list[str]），可为空表示无约束
    required_algo_env_keys = db.Column(db.Text, nullable=True)


class Submission(db.Model):
    __tablename__ = 'submissions'
    id = db.Column(db.Integer, primary_key=True)
    submission_name = db.Column(db.String(120), nullable=False)
    algorithm_image_url = db.Column(db.String(255), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='Pending')  # Pending, Running, Succeeded, Failed, Cancelled
    score = db.Column(db.Float, nullable=True)
    metrics_json = db.Column(db.Text, nullable=True)  # 保存 evaluator 上报的所有指标
    k8s_job_name = db.Column(db.String(120), nullable=True, unique=True)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    user = db.relationship('User', backref=db.backref('submissions', lazy=True))
    leaderboard_id = db.Column(db.Integer, db.ForeignKey('leaderboards.id'), nullable=False)

    # 新增：记录本次任务注入到 algo 容器的 ENV（全部 + 笛卡尔积变量 key 列表）
    algo_env_json = db.Column(db.Text, nullable=True)        # JSON dict[str,str]
    algo_env_grid_keys = db.Column(db.Text, nullable=True)   # JSON list[str]

    __table_args__ = (db.UniqueConstraint('user_id', 'submission_name', name='_user_submission_name_uc'),)


# 新增：积分事件（每次超越 SOTA 的成功提交都会记录）
class PointsEvent(db.Model):
    __tablename__ = 'points_events'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    leaderboard_id = db.Column(db.Integer, db.ForeignKey('leaderboards.id'), nullable=False)
    submission_id = db.Column(db.Integer, db.ForeignKey('submissions.id'), nullable=False)

    year_month = db.Column(db.String(7), nullable=False)  # 形如 "2025-10"
    points = db.Column(db.Float, nullable=False)
    delta = db.Column(db.Float, nullable=False)  # score - sota_score（>=0）
    multiplier = db.Column(db.Float, nullable=False, default=1.0)  # 例如 ≥80 分翻倍 = 2.0
    difficulty_factor = db.Column(db.Float, nullable=False, default=1.0)
    score = db.Column(db.Float, nullable=False)
    sota_score = db.Column(db.Float, nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)


# 新增：月度累计分（不删除历史）
class MonthlyPoints(db.Model):
    __tablename__ = 'monthly_points'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    year_month = db.Column(db.String(7), nullable=False)  # "YYYY-MM"
    total_points = db.Column(db.Float, nullable=False, default=0.0)
    __table_args__ = (db.UniqueConstraint('user_id', 'year_month', name='_user_month_uc'),)


# ============= 提交日志持久化表 =============
class SubmissionLog(db.Model):
    __tablename__ = 'submission_logs'
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('submissions.id'), nullable=False, unique=True)
    evaluator_log = db.Column(db.Text, nullable=True)
    algorithm_log = db.Column(db.Text, nullable=True)
    pod_name = db.Column(db.String(255), nullable=True)
    finalized_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    submission = db.relationship('Submission', backref=db.backref('log_record', uselist=False))

# ============= 新增：逐行日志表（流式采集用） =============
class SubmissionLogLine(db.Model):
    __tablename__ = 'submission_log_lines'
    id            = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('submissions.id'), nullable=False, index=True)
    container     = db.Column(db.String(64), nullable=False)  # 'evaluator-container' | 'submitter-container'
    seq           = db.Column(db.Integer, nullable=False)      # 行号，从 0 开始
    line          = db.Column(db.Text, nullable=False)
    ts            = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('submission_id', 'container', 'seq', name='_sub_container_seq_uc'),
        db.Index('ix_sub_cont_seq', 'submission_id', 'container', 'seq'),
    )


# ============= 新增：题目级评测明细表（用于 Excel） =============
class SubmissionEvalDetail(db.Model):
    __tablename__ = 'submission_eval_details'
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('submissions.id'), nullable=False, index=True)
    leaderboard_id = db.Column(db.Integer, db.ForeignKey('leaderboards.id'), nullable=False, index=True)

    question_id = db.Column(db.String(64), nullable=True)
    question = db.Column(db.Text, nullable=True)
    gold_answer = db.Column(db.Text, nullable=True)
    pred_answer = db.Column(db.Text, nullable=True)
    is_correct = db.Column(db.Boolean, nullable=True)

    latency_ms = db.Column(db.Float, nullable=True)
    used_tokens = db.Column(db.Integer, nullable=True)
    retrieved = db.Column(db.Integer, nullable=True)

    eval_prompt = db.Column(db.Text, nullable=True)   # 评测时使用的 prompt（可选）
    extra_json = db.Column(db.Text, nullable=True)    # 额外字段（算法/榜单自定义 KV）

    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    submission = db.relationship('Submission', backref=db.backref('eval_details', lazy=True))


# ==============================================================
# 补丁 1: 粘贴到 app.py 第 193 行之后（SubmissionEvalDetail 定义之后）
# ==============================================================
class EnvPreset(db.Model):
    __tablename__ = 'env_presets'
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    image_url   = db.Column(db.String(255), nullable=False)
    name        = db.Column(db.String(120), nullable=False)
    env_json    = db.Column(db.Text, nullable=False)
    source      = db.Column(db.String(20), default='manual')
    source_submission_id = db.Column(db.Integer, nullable=True)
    is_default  = db.Column(db.Boolean, default=False)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    user = db.relationship('User', backref=db.backref('env_presets', lazy=True))
    __table_args__ = (
        db.UniqueConstraint('user_id', 'image_url', 'name', name='_user_image_preset_uc'),
    )
# ======================================================================
# ENV 预设 CRUD
# ======================================================================
@app.route('/api/env-presets', methods=['GET'])
@jwt_required()
def list_env_presets():
    user_id = get_jwt_identity()
    image_url = (request.args.get('image_url') or '').strip()
    q = EnvPreset.query.filter_by(user_id=user_id)
    if image_url: q = q.filter_by(image_url=image_url)
    q = q.order_by(EnvPreset.is_default.desc(), EnvPreset.updated_at.desc())
    return jsonify([{"id": p.id, "name": p.name, "image_url": p.image_url,
        "env": json.loads(p.env_json) if p.env_json else {},
        "source": p.source, "is_default": p.is_default,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    } for p in q.all()])

@app.route('/api/env-presets', methods=['POST'])
@jwt_required()
def create_env_preset():
    user_id = get_jwt_identity()
    data = request.get_json() or {}
    image_url = (data.get('image_url') or '').strip()
    name = (data.get('name') or '').strip()
    env_dict = data.get('env') or {}
    if not image_url or not name: return jsonify({"msg": "image_url, name 必填"}), 400
    if not isinstance(env_dict, dict): return jsonify({"msg": "env 必须是 dict"}), 400
    env_dict = _sanitize_env_dict(env_dict)
    set_default = bool(data.get('is_default', False))
    if set_default:
        EnvPreset.query.filter_by(user_id=user_id, image_url=image_url, is_default=True).update({"is_default": False})
    preset = EnvPreset(user_id=user_id, image_url=image_url, name=name,
        env_json=json.dumps(env_dict, ensure_ascii=False), source='manual', is_default=set_default)
    db.session.add(preset)
    try: db.session.commit()
    except IntegrityError:
        db.session.rollback(); return jsonify({"msg": f"预设名「{name}」已存在"}), 409
    return jsonify({"ok": True, "preset_id": preset.id}), 201

@app.route('/api/env-presets/<int:preset_id>', methods=['PUT'])
@jwt_required()
def update_env_preset(preset_id):
    user_id = get_jwt_identity()
    preset = EnvPreset.query.get(preset_id)
    if not preset: return jsonify({"msg": "Not found"}), 404
    if str(preset.user_id) != str(user_id): return jsonify({"msg": "无权"}), 403
    data = request.get_json() or {}
    if 'name' in data and data['name']: preset.name = str(data['name']).strip()
    if 'env' in data and isinstance(data['env'], dict):
        preset.env_json = json.dumps(_sanitize_env_dict(data['env']), ensure_ascii=False)
    if 'is_default' in data:
        if data['is_default']:
            EnvPreset.query.filter(EnvPreset.user_id == user_id, EnvPreset.image_url == preset.image_url,
                EnvPreset.is_default == True, EnvPreset.id != preset.id).update({"is_default": False})
        preset.is_default = bool(data['is_default'])
    try: db.session.commit()
    except IntegrityError: db.session.rollback(); return jsonify({"msg": "名称冲突"}), 409
    return jsonify({"ok": True})

@app.route('/api/env-presets/<int:preset_id>', methods=['DELETE'])
@jwt_required()
def delete_env_preset(preset_id):
    user_id = get_jwt_identity()
    preset = EnvPreset.query.get(preset_id)
    if not preset: return jsonify({"msg": "Not found"}), 404
    if str(preset.user_id) != str(user_id): return jsonify({"msg": "无权"}), 403
    db.session.delete(preset); db.session.commit()
    return jsonify({"ok": True})

@app.route('/api/env-presets/from-submission/<int:sub_id>', methods=['POST'])
@jwt_required()
def create_preset_from_submission(sub_id):
    user_id = get_jwt_identity()
    sub = Submission.query.get(sub_id)
    if not sub: return jsonify({"msg": "Not found"}), 404
    if str(sub.user_id) != str(user_id):
        if get_jwt().get('role') != 'admin': return jsonify({"msg": "只能从自己的提交导入"}), 403
    env_dict = json.loads(sub.algo_env_json) if sub.algo_env_json else {}
    grid_keys = set(json.loads(sub.algo_env_grid_keys)) if sub.algo_env_grid_keys else set()
    fixed = _sanitize_env_dict({k: v for k, v in env_dict.items() if k not in grid_keys})
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip() or f"从提交#{sub_id}导入"
    set_default = bool(data.get('is_default', False))
    if set_default:
        EnvPreset.query.filter_by(user_id=user_id, image_url=sub.algorithm_image_url, is_default=True).update({"is_default": False})
    preset = EnvPreset(user_id=user_id, image_url=sub.algorithm_image_url, name=name,
        env_json=json.dumps(fixed, ensure_ascii=False), source='history',
        source_submission_id=sub.id, is_default=set_default)
    db.session.add(preset)
    try: db.session.commit()
    except IntegrityError: db.session.rollback(); return jsonify({"msg": f"「{name}」已存在"}), 409
    return jsonify({"ok": True, "preset_id": preset.id, "env": fixed}), 201

@app.route('/api/env-presets/history-submissions', methods=['GET'])
@jwt_required()
def list_history_submissions_for_preset():
    user_id = get_jwt_identity()
    image_url = (request.args.get('image_url') or '').strip()
    q = Submission.query.filter_by(user_id=user_id)
    if image_url: q = q.filter_by(algorithm_image_url=image_url)
    q = q.filter(Submission.algo_env_json.isnot(None)).order_by(Submission.submitted_at.desc()).limit(20)
    results = []
    for s in q.all():
        env_dict = json.loads(s.algo_env_json) if s.algo_env_json else {}
        grid_keys = set(json.loads(s.algo_env_grid_keys)) if s.algo_env_grid_keys else set()
        fixed = {k: v for k, v in env_dict.items() if k not in grid_keys}
        results.append({"submission_id": s.id, "submission_name": s.submission_name,
            "status": s.status, "score": s.score,
            "submitted_at": s.submitted_at.isoformat() if s.submitted_at else None,
            "env_preview": fixed, "env_count": len(fixed)})
    return jsonify(results)

@app.route('/api/env-presets/image-history', methods=['GET'])
@jwt_required()
def image_url_history():
    from sqlalchemy import func
    board_id = request.args.get('leaderboard_id', type=int)
    gc = dict(db.session.query(Submission.algorithm_image_url, func.count()).group_by(Submission.algorithm_image_url).all())
    bc = {}
    if board_id:
        bc = dict(db.session.query(Submission.algorithm_image_url, func.count()).filter_by(leaderboard_id=board_id).group_by(Submission.algorithm_image_url).all())
    ranked = sorted(gc.keys(), key=lambda i: (bc.get(i,0)*10000+gc.get(i,0)), reverse=True)
    return jsonify([{"image_url": img, "board_count": bc.get(img,0), "global_count": gc.get(img,0)} for img in ranked[:50]])

@app.route('/api/env-presets/key-history', methods=['GET'])
@jwt_required()
def env_key_value_history():
    board_id = request.args.get('leaderboard_id', type=int)
    if not board_id: return jsonify({})
    agg = {}
    for sub in Submission.query.filter_by(leaderboard_id=board_id).filter(Submission.algo_env_json.isnot(None)).order_by(Submission.submitted_at.desc()).limit(500).all():
        try:
            env = json.loads(sub.algo_env_json)
            if isinstance(env, dict):
                for k, v in env.items():
                    vs = str(v).strip()
                    if vs: agg.setdefault(k, {}); agg[k][vs] = agg[k].get(vs, 0) + 1
        except: pass
    for p in EnvPreset.query.all():
        try:
            env = json.loads(p.env_json)
            if isinstance(env, dict):
                for k, v in env.items():
                    vs = str(v).strip()
                    if vs: agg.setdefault(k, {}); agg[k][vs] = agg[k].get(vs, 0) + 1
        except: pass
    return jsonify({k: [v for v, _ in sorted(vc.items(), key=lambda x: -x[1])[:15]] for k, vc in agg.items()})

# ======================================================================
# 命令行工具
# ======================================================================
@app.cli.command("init-db")
@click.option('--create-defaults', is_flag=True, help='Create default admin, creator, and participant users.')
def init_db_command(create_defaults):
    db.create_all()
    print("Initialized the database.")
    if create_defaults:
        create_default_user('admin', 'adminpass', 'admin')
        create_default_user('l_creator', 'creatorpass', 'creator')
        create_default_user('p_user1', 'user1pass', 'participant')


def create_default_user(username, password, role):
    if not User.query.filter_by(username=username).first():
        user = User(username=username, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        print(f"Created default {role} user: {username}")
    else:
        print(f"User '{username}' already exists.")


@app.cli.command("create-admin")
@click.argument("username")
@click.argument("password")
def create_admin_command(username, password):
    user = User.query.filter_by(username=username).first()
    if user:
        print(f"Error: User '{username}' already exists.")
        return
    new_admin = User(username=username, role='admin')
    new_admin.set_password(password)
    db.session.add(new_admin)
    db.session.commit()
    print(f"Admin user '{username}' created successfully.")


# ============= 批量回收固化日志的命令 =============
@app.cli.command("harvest-logs")
@click.option('--only-missing', is_flag=True, help='Only harvest submissions without persisted logs.')
def harvest_logs_command(only_missing):
    """
    回收所有（或仅缺失）终态任务的 Pod 日志并固化。
    """
    if not k8s_core_v1 or not k8s_batch_v1:
        print("Kubernetes client not available.")
        return

    q = Submission.query.filter(Submission.status.in_(('Succeeded', 'Failed', 'Cancelled')))
    if only_missing:
        q = q.outerjoin(SubmissionLog, SubmissionLog.submission_id == Submission.id).filter(SubmissionLog.id.is_(None))
    subs = q.all()
    print(f"Harvesting logs for {len(subs)} submissions...")
    ok = 0
    for sub in subs:
        try:
            changed = _persist_submission_logs(sub)
            if changed:
                ok += 1
        except Exception as e:
            print(f"  - sub {sub.id} harvest failed: {e}")
    print(f"Done. Persisted {ok} submissions.")


# ======================================================================
# 认证装饰器
# ======================================================================
def admin_required():
    def wrapper(fn):
        @wraps(fn)
        @jwt_required()
        def decorator(*args, **kwargs):
            current_user_id = get_jwt_identity()
            user = User.query.get(current_user_id)
            if not user or user.role != 'admin':
                return jsonify({"msg": "Admins only access!"}), 403
            return fn(*args, **kwargs)
        return decorator
    return wrapper


def creator_required():
    def wrapper(fn):
        @wraps(fn)
        @jwt_required()
        def decorator(*args, **kwargs):
            current_user_id = get_jwt_identity()
            user = User.query.get(current_user_id)
            if not user or user.role not in ['admin', 'creator']:
                return jsonify({"msg": "Admins or Creators only access!"}), 403
            return fn(*args, **kwargs)
        return decorator
    return wrapper


def is_admin_or_owner(owner_id):
    current_user_id = get_jwt_identity()
    if current_user_id == str(owner_id):
        return True
    claims = get_jwt()
    if claims.get('role') == 'admin':
        return True
    return False


# ======================================================================
# 工具：ENV 解析与注入
# ======================================================================

_ENV_KEY_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
_RESERVED_ENV = {"SUBMISSION_ID", "LEADERBOARD_ID", "API_INTERNAL_URL", "ALGO_API_ENDPOINT"}


def _parse_env_text(text: str | None) -> dict:
    """
    解析多行 KEY=VALUE，忽略空行与注释；前后空白清理；不校验重复（后续合并时处理）。
    """
    out = {}
    if not text:
        return out
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            # 忽略非法行
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        if not k:
            continue
        out[k] = v
    return out


def _sanitize_env_dict(d: dict | None) -> dict:
    """
    过滤非法变量名；确保值转为字符串。
    """
    if not isinstance(d, dict):
        return {}
    cleaned = {}
    for k, v in d.items():
        if not isinstance(k, str):
            continue
        key = k.strip()
        if not key:
            continue
        if not _ENV_KEY_RE.match(key):
            # 非法变量名跳过
            continue
        cleaned[key] = "" if v is None else str(v)
    return cleaned


def _merge_envs(fixed_env: dict, grid_env: dict) -> dict:
    """
    合并 ENV：网格参数 > 固定文本；保留键名大小写。
    """
    base = dict(fixed_env or {})
    base.update(grid_env or {})
    return base


def _dict_to_envvars(extra_env: dict) -> list:
    """
    将 dict 转为 V1EnvVar 列表；跳过保留键（避免用户覆盖内置关键变量）。
    """
    envs = []
    for k, v in (extra_env or {}).items():
        if k in _RESERVED_ENV:
            continue
        envs.append(client.V1EnvVar(name=k, value=str(v)))
    return envs


def _normalize_required_env_keys(val) -> str | None:
    """
    将创建 / 编辑榜单时传入的 required_algo_env_keys 统一转成 JSON list[str] 存表。
    支持：
      - list[str]
      - 逗号/换行分隔的字符串
    返回 JSON 字符串或 None（表示无约束）。
    """
    if not val:
        return None
    keys: list[str] = []
    if isinstance(val, str):
        raw = val.replace("\n", ",")
        for seg in raw.split(","):
            s = seg.strip()
            if s:
                keys.append(s)
    elif isinstance(val, (list, tuple, set)):
        for x in val:
            if x is None:
                continue
            s = str(x).strip()
            if s:
                keys.append(s)
    else:
        return None
    if not keys:
        return None
    uniq = sorted(set(keys))
    return json.dumps(uniq, ensure_ascii=False)


def _get_leaderboard_required_env_keys(board: Leaderboard) -> list[str]:
    """
    读取榜单级必填 ENV key。
    - 优先使用 Leaderboard.required_algo_env_keys（JSON list[str]）
    - 若为空，则退回全局环境变量 LB_REQUIRED_ALGO_ENV_KEYS（逗号/空白分隔）
    """
    if board and board.required_algo_env_keys:
        try:
            arr = json.loads(board.required_algo_env_keys)
            if isinstance(arr, list):
                out = []
                for v in arr:
                    s = str(v).strip()
                    if s:
                        out.append(s)
                if out:
                    return out
        except Exception:
            pass

    raw = os.environ.get("LB_REQUIRED_ALGO_ENV_KEYS", "")
    if not raw:
        return []
    keys: list[str] = []
    for seg in re.split(r"[,\s]+", raw):
        s = seg.strip()
        if s:
            keys.append(s)
    return keys


# ======================================================================
# 模块 1: 管理模块 (Admin & Auth)
# ======================================================================
@app.route('/api/login', methods=['POST'])
def login():
    username = request.json.get('username', None)
    password = request.json.get('password', None)
    user = User.query.filter_by(username=username).first()
    if user and user.check_password(password):
        additional_claims = {"role": user.role}
        access_token = create_access_token(identity=str(user.id), additional_claims=additional_claims)
        return jsonify(access_token=access_token)
    return jsonify({"msg": "Bad username or password"}), 401


@app.route('/api/admin/users', methods=['POST'])
@admin_required()
def create_user():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'participant')
    if role not in ['admin', 'creator', 'participant']:
        return jsonify({"msg": "Invalid role specified"}), 400
    if not username or not password:
        return jsonify({"msg": "Missing username or password"}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({"msg": "Username already exists"}), 409
    new_user = User(username=username, role=role)
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    return jsonify({"id": new_user.id, "username": new_user.username, "role": new_user.role}), 201


@app.route('/api/admin/users', methods=['GET'])
@admin_required()
def get_users():
    users = User.query.all()
    user_list = [
        {"id": u.id, "username": u.username, "role": u.role, "created_at": u.created_at.isoformat()}
        for u in users
    ]
    return jsonify(user_list)


@app.route('/api/admin/users/<int:user_id>/password', methods=['PUT'])
@admin_required()
def update_user_password(user_id):
    user = User.query.get_or_404(user_id)
    password = request.json.get('password')
    if not password:
        return jsonify({"msg": "Missing password"}), 400
    user.set_password(password)
    db.session.commit()
    return jsonify({"msg": f"Password for user {user.username} updated."})


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required()
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    current_user_id = get_jwt_identity()
    if str(user.id) == current_user_id:
        return jsonify({"msg": "Cannot delete yourself"}), 403
    if user.role == 'admin':
        admin_count = User.query.filter_by(role='admin').count()
        if admin_count <= 1:
            return jsonify({"msg": "Cannot delete the last admin"}), 403
    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({"msg": f"User {user.username} deleted."})
    except IntegrityError:
        db.session.rollback()
        return jsonify({"msg": "Could not delete user. Check for associated leaderboards or submissions if cascade delete is not set."}), 409


# ======================================================================
# 模块 2: 榜单发布模块 (Leaderboard Management)
# ======================================================================
@app.route('/api/leaderboards', methods=['POST'])
@creator_required()
def create_leaderboard():
    data = request.get_json()
    owner_id = get_jwt_identity()
    try:
        resource_spec_dict = data.get('resource_spec', {})
        resource_spec_json = json.dumps(resource_spec_dict)
    except (TypeError, json.JSONDecodeError):
        return jsonify({"msg": "Invalid resource_spec format. Must be a JSON object."}), 400

    required_env_json = _normalize_required_env_keys(data.get('required_algo_env_keys'))

    new_board = Leaderboard(
        name=data.get('name'),
        description=data.get('description'),
        evaluator_image=data.get('evaluator_image'),
        baseline_image=data.get('baseline_image'),
        resource_spec=resource_spec_json,
        version=data.get('version'),
        owner_id=int(owner_id),
        difficulty_factor=float(data.get('difficulty_factor', 1.0)),
        sota_score=(float(data['sota_score']) if data.get('sota_score') is not None else None),
        required_algo_env_keys=required_env_json
    )
    if not new_board.name or not new_board.evaluator_image or not new_board.version:
        return jsonify({"msg": "Missing required fields: name, evaluator_image, version"}), 400

    db.session.add(new_board)
    db.session.commit()
    return jsonify({"id": new_board.id, "name": new_board.name}), 201


@app.route('/api/leaderboards/<int:leaderboard_id>', methods=['PUT'])
@creator_required()
def update_leaderboard(leaderboard_id):
    board = Leaderboard.query.get_or_404(leaderboard_id)
    if not is_admin_or_owner(board.owner_id):
        return jsonify({"msg": "Forbidden: You are not the owner or an admin"}), 403

    data = request.get_json()
    updated = False
    if 'name' in data and data['name']:
        board.name = data['name']; updated = True
    if 'description' in data:
        board.description = data['description']; updated = True
    if 'evaluator_image' in data and data['evaluator_image']:
        board.evaluator_image = data['evaluator_image']; updated = True
    if 'baseline_image' in data:
        board.baseline_image = data['baseline_image']; updated = True
    if 'resource_spec' in data:
        try:
            board.resource_spec = json.dumps(data['resource_spec']); updated = True
        except (TypeError, json.JSONDecodeError):
            return jsonify({"msg": "Invalid resource_spec format. Must be a JSON object."}), 400
    if 'version' in data and data['version']:
        board.version = data['version']; updated = True
    if 'difficulty_factor' in data:
        try:
            board.difficulty_factor = float(data['difficulty_factor']); updated = True
        except (ValueError, TypeError):
            return jsonify({"msg": "difficulty_factor must be a number"}), 400
    if 'sota_score' in data:
        try:
            board.sota_score = (float(data['sota_score']) if data['sota_score'] is not None else None); updated = True
        except (ValueError, TypeError):
            return jsonify({"msg": "sota_score must be a number or null"}), 400
    if 'required_algo_env_keys' in data:
        board.required_algo_env_keys = _normalize_required_env_keys(data.get('required_algo_env_keys'))
        updated = True

    if not updated:
        return jsonify({"msg": "No valid fields provided for update"}), 400

    db.session.commit()
    return jsonify({
        "msg": f"Leaderboard {board.name} updated.",
        "id": board.id,
        "name": board.name,
        "description": board.description,
        "evaluator_image": board.evaluator_image,
        "baseline_image": board.baseline_image,
        "resource_spec": json.loads(board.resource_spec),
        "version": board.version,
        "difficulty_factor": board.difficulty_factor,
        "sota_score": board.sota_score,
        "required_algo_env_keys": (
            json.loads(board.required_algo_env_keys) if board.required_algo_env_keys else None
        )
    })


@app.route('/api/leaderboards/manage', methods=['GET'])
@creator_required()
def get_my_leaderboards():
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    if user.role == 'admin':
        boards = Leaderboard.query.order_by(Leaderboard.name).all()
    else:
        boards = Leaderboard.query.filter_by(owner_id=user_id).order_by(Leaderboard.name).all()

    board_list = [
        {
            "id": b.id,
            "name": b.name,
            "version": b.version,
            "description": b.description,
            "owner_username": b.owner.username,
            "difficulty_factor": b.difficulty_factor,
            "sota_score": b.sota_score,
            "required_algo_env_keys": (
                json.loads(b.required_algo_env_keys) if b.required_algo_env_keys else None
            )
        } for b in boards
    ]
    return jsonify(board_list)


# ======================================================================
# 模块 3: 打榜模块 (Submission & Tracking)
# ======================================================================
# ======================================================================
# 本机（无 K8s）执行后备：当集群不可用或显式开启 LOCAL_EXECUTOR 时，
# 不创建 K8s Job，而是在本机以子进程方式拉起 evaluator，由 evaluator
# 通过 /api/internal/submission/<id>/score 回传结果（与 K8s 路径同契约）。
# K8s 相关代码与代理设置保持原样，互不影响。
# ======================================================================
import threading as _threading
import subprocess as _subprocess


def _local_exec_enabled() -> bool:
    if os.environ.get("LOCAL_EXECUTOR", "").strip() in ("1", "true", "True"):
        return True
    return k8s_batch_v1 is None


def _local_runner_log(submission_id: int, container: str, text: str):
    """把本机 evaluator 的输出按行写入 SubmissionLogLine，前端可实时查看。"""
    try:
        with app.app_context():
            base = SubmissionLogLine.query.filter_by(
                submission_id=submission_id, container=container
            ).count()
            seq = base
            for line in text.splitlines():
                db.session.add(SubmissionLogLine(
                    submission_id=submission_id, container=container,
                    seq=seq, line=line, ts=datetime.utcnow()))
                seq += 1
            db.session.commit()
    except Exception as e:
        app.logger.error(f"local log persist failed sub {submission_id}: {e}")


def _local_exec_worker(submission_id: int, job_name: str, extra_env: dict):
    # Default to a repo-relative path (backend/ -> repo root) so a fresh clone
    # works without a host-specific path; override via LOCAL_EXEC_CMD.
    _repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cmd = os.environ.get(
        "LOCAL_EXEC_CMD",
        os.path.join(
            _repo_root,
            "leadeboard_apps/spreadsheets_bench/skillopt_eval/run_submission.sh",
        ),
    )
    env = dict(os.environ)
    env["SUBMISSION_ID"] = str(submission_id)
    env.setdefault("API_INTERNAL_URL", os.environ.get("API_INTERNAL_URL", "http://localhost:7789"))
    for k, v in (extra_env or {}).items():
        if v is not None:
            env[str(k)] = str(v)
    app.logger.info(f"[local-exec] sub {submission_id} launching: {cmd} (SKILL_ALGO={env.get('SKILL_ALGO')})")
    captured = []
    try:
        proc = _subprocess.Popen(
            ["bash", cmd], env=env,
            stdout=_subprocess.PIPE, stderr=_subprocess.STDOUT,
            bufsize=1, universal_newlines=True,
        )
        buf = []
        for line in proc.stdout:
            captured.append(line)
            buf.append(line.rstrip("\n"))
            if len(buf) >= 20:
                _local_runner_log(submission_id, "evaluator-container", "\n".join(buf))
                buf = []
        if buf:
            _local_runner_log(submission_id, "evaluator-container", "\n".join(buf))
        rc = proc.wait()
        app.logger.info(f"[local-exec] sub {submission_id} finished rc={rc}")
    except Exception as e:
        app.logger.error(f"[local-exec] sub {submission_id} crashed: {e}")
        captured.append(f"\n[local-exec] crashed: {e}\n")
        rc = -1
    # 固化整段日志 + 兜底状态（evaluator 正常会自行 POST 终态）
    try:
        with app.app_context():
            sub = Submission.query.get(submission_id)
            if sub:
                if not SubmissionLog.query.filter_by(submission_id=submission_id).first():
                    db.session.add(SubmissionLog(
                        submission_id=submission_id,
                        evaluator_log="".join(captured)[-200000:],
                        algorithm_log="(local executor: skill artifact, no algo container)",
                        pod_name=job_name, finalized_at=datetime.utcnow()))
                if rc != 0 and sub.status in ("Submitted", "Pending", "Running"):
                    sub.status = "Failed"
                db.session.commit()
    except Exception as e:
        app.logger.error(f"[local-exec] finalize failed sub {submission_id}: {e}")


def _start_local_job(submission: Submission, extra_env: dict | None = None):
    job_name = f"local-sub-{submission.id}-{uuid.uuid4().hex[:6]}"
    submission.k8s_job_name = job_name
    submission.status = "Running"
    db.session.commit()
    t = _threading.Thread(
        target=_local_exec_worker,
        args=(submission.id, job_name, dict(extra_env or {})),
        daemon=True,
    )
    t.start()
    app.logger.info(f"[local-exec] started local job {job_name} for submission {submission.id}")
    return job_name


def _start_k8s_job(submission: Submission, extra_env: dict | None = None):
    if _local_exec_enabled():
        return _start_local_job(submission, extra_env)
    if not k8s_batch_v1:
        raise ConnectionError("Kubernetes API client not initialized.")
    leaderboard = submission.leaderboard
    if not leaderboard:
        raise ValueError(f"Leaderboard with ID {submission.leaderboard_id} not found for submission {submission.id}")

    job_name = f"eval-sub-{submission.id}-{uuid.uuid4().hex[:6]}"
    submission.k8s_job_name = job_name

    try:
        resource_spec = json.loads(leaderboard.resource_spec)
        job_manifest = create_job_manifest(
            job_name=job_name,
            leaderboard_id=submission.leaderboard_id,
            evaluator_image=leaderboard.evaluator_image,
            algorithm_image=submission.algorithm_image_url,
            resource_spec=resource_spec,
            submission_id=submission.id,
            extra_env=extra_env or {}
        )
        k8s_batch_v1.create_namespaced_job(body=job_manifest, namespace=K8S_NAMESPACE)
        submission.status = 'Pending'
        db.session.commit()
        app.logger.info(f"Successfully created K8s Job: {job_name} for submission {submission.id}")
        return job_name
    except ApiException as e:
        app.logger.error(f"K8s API Error creating job for sub {submission.id}: {e.body}")
        submission.status = 'Failed'
        submission.k8s_job_name = None
        db.session.commit()
        raise e
    except Exception as e:
        app.logger.error(f"Unexpected Error creating job for sub {submission.id}: {e}")
        submission.status = 'Failed'
        submission.k8s_job_name = None
        db.session.commit()
        raise e


@app.route('/api/leaderboard/<int:leaderboard_id>/submit', methods=['POST'])
@jwt_required()
def submit_to_leaderboard(leaderboard_id):
    board = Leaderboard.query.get(leaderboard_id)
    if not board:
        return jsonify({"msg": f"Leaderboard {leaderboard_id} not found"}), 404

    user_id = get_jwt_identity()
    data = request.get_json()
    submission_name = data.get('submission_name')
    algo_image = data.get('algorithm_image_url')
    if not submission_name or not algo_image:
        return jsonify({"msg": "Missing submission_name or algorithm_image_url"}), 400

    # 解析前端新增的 ENV & 网格参数
    fixed_env_raw = data.get('env_text', '') or ''
    grid_params = data.get('params', {}) or {}

    fixed_env = _sanitize_env_dict(_parse_env_text(fixed_env_raw))
    grid_env = _sanitize_env_dict(grid_params) if isinstance(grid_params, dict) else {}

    # 合并：网格 > 固定
    extra_env = _merge_envs(fixed_env, grid_env)

    # 校验必填 ENV（榜单级 + 可选全局环境变量）
    required_keys = _get_leaderboard_required_env_keys(board)
    if required_keys:
        missing = []
        for k in required_keys:
            v = extra_env.get(k)
            if v is None or str(v).strip() == "":
                missing.append(k)
        if missing:
            return jsonify({
                "msg": "Missing required algo env variables",
                "required_keys": required_keys,
                "missing_keys": missing
            }), 400

    existing_submission = Submission.query.filter_by(user_id=user_id, submission_name=submission_name).first()
    if existing_submission:
        return jsonify({"msg": "Submission name already used by this user."}), 409

    # 记录本次任务的 ENV 快照和笛卡尔积变量 key 集合
    algo_env_json = json.dumps(extra_env, ensure_ascii=False) if extra_env else None
    grid_keys = sorted(grid_env.keys()) if grid_env else []
    algo_env_grid_keys = json.dumps(grid_keys, ensure_ascii=False) if grid_keys else None

    new_submission = Submission(
        submission_name=submission_name,
        algorithm_image_url=algo_image,
        status='Submitted',
        user_id=user_id,
        leaderboard_id=leaderboard_id,
        algo_env_json=algo_env_json,
        algo_env_grid_keys=algo_env_grid_keys
    )
    db.session.add(new_submission)
    try:
        db.session.flush()
        job_name = _start_k8s_job(new_submission, extra_env=extra_env)
        db.session.commit()
        return jsonify({
            "message": "Submission received. Evaluation job started.",
            "submission_id": new_submission.id,
            "job_name": job_name,
            "status": new_submission.status
        }), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"msg": "Submission name already used (concurrent request?)."}), 409
    except (ApiException, Exception) as e:
        db.session.rollback()
        error_detail = getattr(e, 'body', str(e))
        return jsonify({"error": "Failed to create K8s job", "details": error_detail}), 500


def _read_pod_log_text(pod_name, container):
    """read_namespaced_pod_log 在本版本 client 下返回 bytes；解码为 UTF-8 文本，
    避免被当作 bytes 存入 Text 列后变成 b'...' 的 repr 乱码。"""
    out = k8s_core_v1.read_namespaced_pod_log(name=pod_name, namespace=K8S_NAMESPACE, container=container)
    if isinstance(out, (bytes, bytearray)):
        out = bytes(out).decode("utf-8", errors="replace")
    return out


def _persist_submission_logs(submission: Submission) -> bool:
    """
    若当前 submission 为终态，尝试从 K8s 读取 evaluator/algorithm 日志并固化到 DB。
    返回 True 表示有写入（新建或更新），False 表示无变化或无法获取。
    """
    if not submission or not submission.k8s_job_name:
        return False
    if submission.status not in ('Succeeded', 'Failed', 'Cancelled'):
        return False
    if not k8s_core_v1:
        app.logger.warning("K8s client not available for persisting logs.")
        return False

    # 已有固化记录则跳过
    existing = SubmissionLog.query.filter_by(submission_id=submission.id).first()
    if existing and existing.evaluator_log is not None and existing.algorithm_log is not None:
        return False

    pod_name = None
    evaluator_log = ""
    algorithm_log = ""

    try:
        pod_list = k8s_core_v1.list_namespaced_pod(
            namespace=K8S_NAMESPACE,
            label_selector=f"job-name={submission.k8s_job_name}"
        )
        if not pod_list.items:
            # 尝试抓取 Job 事件作为替代
            try:
                job_events = k8s_core_v1.list_namespaced_event(
                    namespace=K8S_NAMESPACE,
                    field_selector=f"involvedObject.kind=Job,involvedObject.name={submission.k8s_job_name}"
                )
                event_log = "\n".join([f"{e.last_timestamp.isoformat()}: {e.reason} - {e.message}" for e in job_events.items])
                evaluator_log = f"[No Pod found. Job events]\n{event_log}"
                algorithm_log = ""
                pod_name = None
            except ApiException:
                evaluator_log = "[No Pod found and failed to read Job events.]"
                algorithm_log = ""
                pod_name = None
        else:
            pod = pod_list.items[0]
            pod_name = pod.metadata.name
            # evaluator
            try:
                evaluator_log = _read_pod_log_text(pod_name, "evaluator-container")
            except ApiException as log_e:
                evaluator_log = f"Error retrieving evaluator log: {getattr(log_e, 'body', str(log_e))}"
            # algorithm
            try:
                algorithm_log = _read_pod_log_text(pod_name, "submitter-container")
            except ApiException as log_e:
                algorithm_log = f"Error retrieving algorithm log: {getattr(log_e, 'body', str(log_e))}"

    except ApiException as e:
        app.logger.error(f"Persist logs failed for sub {submission.id}: {e}")
        return False
    except Exception as e:
        app.logger.error(f"Unexpected error persist logs for sub {submission.id}: {e}")
        return False

    # 入库（新建或更新）
    try:
        if existing:
            existing.evaluator_log = evaluator_log
            existing.algorithm_log = algorithm_log
            existing.pod_name = pod_name
            existing.finalized_at = datetime.utcnow()
        else:
            rec = SubmissionLog(
                submission_id=submission.id,
                evaluator_log=evaluator_log,
                algorithm_log=algorithm_log,
                pod_name=pod_name,
                finalized_at=datetime.utcnow()
            )
            db.session.add(rec)
        db.session.commit()
        app.logger.info(f"Persisted logs for submission {submission.id} (pod={pod_name})")
        return True
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"DB error persisting logs for sub {submission.id}: {e}")
        return False


@app.route('/api/submission/<int:sub_id>/logs/download', methods=['GET'])
@jwt_required()
def download_submission_logs(sub_id):
    submission = Submission.query.get_or_404(sub_id)
    if not is_admin_or_owner(submission.user_id):
        return jsonify({"msg": "Forbidden"}), 403

    container = request.args.get('container', 'evaluator-container')
    if container not in ('evaluator-container', 'submitter-container'):
        return jsonify({"msg": "Invalid container name"}), 400

    # ── 路径 A：新表逐行日志 ──
    lines = (
        SubmissionLogLine.query
        .filter_by(submission_id=sub_id, container=container)
        .order_by(SubmissionLogLine.seq)
        .all()
    )
    if lines:
        content = "\n".join(l.line for l in lines)
    else:
        # ── 路径 B：降级到旧 SubmissionLog 表 ──
        log_rec = SubmissionLog.query.filter_by(submission_id=sub_id).first()
        if log_rec:
            content = (log_rec.evaluator_log or "") if container == 'evaluator-container' \
                      else (log_rec.algorithm_log or "")
        else:
            content = ""

    # 文件名：sub{id}_{eval|algo}_{job_name}_{状态}_{UTC时间}.log
    container_tag = 'eval' if container == 'evaluator-container' else 'algo'
    job_tag = (submission.k8s_job_name or 'nojob').replace('/', '-')[:40]
    status_tag = submission.status.lower()
    ts_tag = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    filename = f"sub{sub_id}_{container_tag}_{job_tag}_{status_tag}_{ts_tag}.log"

    from flask import make_response
    resp = make_response(content)
    resp.headers['Content-Type']        = 'text/plain; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp.headers['X-Filename']          = filename
    return resp


def _sync_submission_status(submission: Submission):
    if not k8s_batch_v1:
        app.logger.warning("K8s client not available, skipping status sync.")
        return submission.status

    original_status = submission.status

    if submission.status in ['Pending', 'Running'] and submission.k8s_job_name:
        try:
            job_status_resp = k8s_batch_v1.read_namespaced_job_status(
                name=submission.k8s_job_name, namespace=K8S_NAMESPACE
            )
            k8s_stat = job_status_resp.status
            new_status = submission.status
            if k8s_stat.succeeded and k8s_stat.succeeded > 0:
                new_status = 'Succeeded'
            elif k8s_stat.failed and k8s_stat.failed > 0:
                new_status = 'Failed'
            elif k8s_stat.active and k8s_stat.active > 0:
                new_status = 'Running'

            if new_status != submission.status:
                submission.status = new_status
                db.session.commit()
                app.logger.info(f"Synced submission {submission.id} status to {new_status} based on K8s job {submission.k8s_job_name}")

                # 状态变为 Succeeded 时，若 score 还没写但 metrics 有 score 字段，则补填
                if new_status == 'Succeeded':
                    has_score = submission.score is not None
                    try:
                        mj = json.loads(submission.metrics_json) if getattr(submission, 'metrics_json', None) else None
                    except Exception:
                        mj = None
                    if not has_score and isinstance(mj, dict) and ('score' in mj) and (mj['score'] is not None):
                        try:
                            submission.score = float(mj['score'])
                            db.session.commit()
                        except Exception:
                            db.session.rollback()
                    # 再尝试计分
                    try:
                        _create_points_event_if_eligible(submission)
                    except Exception as _pe:
                        app.logger.error(f"Post-sync points creation failed for sub {submission.id}: {_pe}")

        except ApiException as e:
            if e.status == 404:
                app.logger.warning(f"Job {submission.k8s_job_name} not found in K8s for sub {submission.id}, marking as Failed.")
                submission.status = 'Failed'
                db.session.commit()
            else:
                app.logger.error(f"K8s error checking job {submission.k8s_job_name} for sub {submission.id}: {e}")
        except Exception as e:
            app.logger.error(f"Unexpected error syncing status for sub {submission.id}: {e}")

    # 若进入终态，尝试固化日志（只要能拿到）
    if submission.status in ('Succeeded', 'Failed', 'Cancelled') and original_status != submission.status:
        try:
            _persist_submission_logs(submission)
        except Exception as e:
            app.logger.error(f"Persist logs on terminal transition failed for sub {submission.id}: {e}")

    return submission.status


@app.route('/api/my-submissions', methods=['GET'])
@jwt_required()
def get_my_submissions():
    user_id = get_jwt_identity()

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)

    # 1. 基础查询 + 字段优化 (保持不变)
    query = Submission.query.filter_by(user_id=user_id).order_by(Submission.submitted_at.desc())
    query = query.options(load_only(
        Submission.id,
        Submission.submission_name,
        Submission.leaderboard_id,
        Submission.status,
        Submission.score,
        Submission.submitted_at
        # metrics_json 依然不加载，那是为了详情页留的
    ))

    # 2. 分页 (保持不变)
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    submission_list = []
    for sub in pagination.items:
        # =======================================================
        # 🚨 核心修复：把状态同步加回来！
        # 修改：不再过滤状态，当前页的 10 条数据全部尝试同步。
        # 既然只有 10 条，全量同步的开销是可以接受的，这样能确保状态绝对准确。
        # =======================================================
        try:
            # 调用你的 K8s 同步函数
            _sync_submission_status(sub)

            # 【关键】因为 _sync 可能会 commit 事务，
            # 我们需要刷新当前对象以获取最新状态
            db.session.commit()  # 确保状态入库
            db.session.refresh(sub)
        except Exception as e:
            # 即使同步出错（比如 K8s API 偶尔抖动），只记录日志，不要崩掉整个列表
            app.logger.error(f"Sync error for {sub.id}: {e}")
        # =======================================================

        submission_list.append({
            "id": sub.id,
            "name": sub.submission_name,
            "leaderboard_id": sub.leaderboard_id,
            "leaderboard_name": sub.leaderboard.name if sub.leaderboard else str(sub.leaderboard_id),
            "status": sub.status,  # 这里现在拿到的是最新状态了
            "score": sub.score,
            "metrics": None,
            "submitted_at": sub.submitted_at.isoformat()
        })

    return jsonify({
        "items": submission_list,
        "total": pagination.total,
        "pages": pagination.pages,
        "current_page": page,
        "per_page": per_page
    })


@app.route('/api/submission/<int:sub_id>/logs', methods=['GET'])
@jwt_required()
def get_submission_logs(sub_id):
    submission = Submission.query.get_or_404(sub_id)
    if not is_admin_or_owner(submission.user_id):
        return jsonify({"msg": "Forbidden: You are not the owner or an admin"}), 403

    eval_offset = int(request.args.get('eval_offset', 0))
    algo_offset = int(request.args.get('algo_offset', 0))
    limit       = min(int(request.args.get('limit', 500)), 2000)
    finalized   = submission.status in ('Succeeded', 'Failed', 'Cancelled')

    # ── 路径 A：新表 SubmissionLogLine 有数据（流式采集写入的） ──
    eval_count = SubmissionLogLine.query.filter_by(
        submission_id=sub_id, container='evaluator-container'
    ).count()
    algo_count = SubmissionLogLine.query.filter_by(
        submission_id=sub_id, container='submitter-container'
    ).count()

    if eval_count > 0 or algo_count > 0:
        eval_rows = (
            SubmissionLogLine.query
            .filter_by(submission_id=sub_id, container='evaluator-container')
            .filter(SubmissionLogLine.seq >= eval_offset)
            .order_by(SubmissionLogLine.seq)
            .limit(limit)
            .all()
        )
        algo_rows = (
            SubmissionLogLine.query
            .filter_by(submission_id=sub_id, container='submitter-container')
            .filter(SubmissionLogLine.seq >= algo_offset)
            .order_by(SubmissionLogLine.seq)
            .limit(limit)
            .all()
        )
        return jsonify({
            "mode": "incremental",
            "evaluator_lines": [r.line for r in eval_rows],
            "algorithm_lines": [r.line for r in algo_rows],
            "eval_next_offset": eval_rows[-1].seq + 1 if eval_rows else eval_offset,
            "algo_next_offset": algo_rows[-1].seq + 1 if algo_rows else algo_offset,
            "eval_total":  eval_count,
            "algo_total":  algo_count,
            "finalized":   finalized,
        })

    # ── 路径 B：降级到旧 SubmissionLog 表（历史存量数据） ──
    log_rec = SubmissionLog.query.filter_by(submission_id=submission.id).first()
    if log_rec and (log_rec.evaluator_log is not None or log_rec.algorithm_log is not None):
        return jsonify({
            "mode": "legacy",
            "evaluator_log": log_rec.evaluator_log or "",
            "algorithm_log": log_rec.algorithm_log or "",
            "finalized": finalized,
        })

    # ── 路径 C：两张表都没有，直接从 K8s 实时拉取（兜底） ──
    if not submission.k8s_job_name:
        return jsonify({"msg": "Job has not started or has no K8s name"}), 404
    if not k8s_core_v1:
        return jsonify({"msg": "Kubernetes API client not available"}), 503

    try:
        pod_list = k8s_core_v1.list_namespaced_pod(
            namespace=K8S_NAMESPACE,
            label_selector=f"job-name={submission.k8s_job_name}"
        )
        if not pod_list.items:
            try:
                job_events = k8s_core_v1.list_namespaced_event(
                    namespace=K8S_NAMESPACE,
                    field_selector=f"involvedObject.kind=Job,involvedObject.name={submission.k8s_job_name}"
                )
                event_log = "\n".join([
                    f"{e.last_timestamp.isoformat()}: {e.reason} - {e.message}"
                    for e in job_events.items
                ])
                evaluator_log = f"Pod not found (possibly cleaned up). Job events:\n{event_log}"
                algorithm_log = ""
            except ApiException:
                return jsonify({"msg": "Pod not found for job and failed to get job events"}), 404
        else:
            pod_name = pod_list.items[0].metadata.name
            try:
                evaluator_log = _read_pod_log_text(pod_name, "evaluator-container")
            except ApiException as log_e:
                evaluator_log = f"Error retrieving evaluator log: {log_e.body}"
            try:
                algorithm_log = _read_pod_log_text(pod_name, "submitter-container")
            except ApiException as log_e:
                algorithm_log = f"Error retrieving algorithm log: {log_e.body}"

        if finalized:
            try:
                _persist_submission_logs(submission)
            except Exception as e:
                app.logger.error(f"Persist logs after GET failed for sub {submission.id}: {e}")

        return jsonify({
            "mode": "legacy",
            "evaluator_log": evaluator_log,
            "algorithm_log": algorithm_log,
            "finalized": finalized,
        })

    except ApiException as e:
        app.logger.error(f"Error getting logs for job {submission.k8s_job_name}: {e}")
        return jsonify({"msg": "Error retrieving logs", "details": e.body}), 500


@app.route('/api/submission/<int:sub_id>/rerun', methods=['POST'])
@jwt_required()
def rerun_submission(sub_id):
    original_submission = Submission.query.get_or_404(sub_id)
    user_id = get_jwt_identity()
    if str(original_submission.user_id) != user_id:
        return jsonify({"msg": "Forbidden: You can only rerun your own submissions"}), 403

    rerun_count = Submission.query.filter(
        Submission.user_id == user_id,
        Submission.submission_name.like(f"{original_submission.submission_name} (rerun %)")
    ).count()
    new_submission_name = f"{original_submission.submission_name} (rerun {rerun_count + 1})"
    if Submission.query.filter_by(user_id=user_id, submission_name=new_submission_name).first():
        return jsonify({"msg": "Generated rerun name already exists. Please rename original or try again."}), 409

    new_submission = Submission(
        submission_name=new_submission_name,
        algorithm_image_url=original_submission.algorithm_image_url,
        status='Submitted',
        user_id=user_id,
        leaderboard_id=original_submission.leaderboard_id,
        algo_env_json=original_submission.algo_env_json,
        algo_env_grid_keys=original_submission.algo_env_grid_keys
    )
    db.session.add(new_submission)

    try:
        db.session.flush()

        # 判断原任务的提交类型：upload-mode（文件上传打榜）vs 镜像打榜
        is_upload_mode = (original_submission.algorithm_image_url == "upload-mode")

        if is_upload_mode:
            # ========== upload-mode rerun ==========

            def _recover_upload_qa_file(submission_id: int) -> str | None:
                """
                存量兼容：当 algo_env_json 里没有记录文件路径时（历史任务），
                根据文件命名规则 sub_{id}_{ts}_{name} 在磁盘上反查。
                两种存储模式都支持：
                  - hostPath：扫 /models/uploads/，返回容器内可见的 /models/uploads/{basename}
                  - PVC：扫 UPLOADS_HOST_PATH，返回 http://.../internal/uploads/{basename}
                若找到多个（同一 sub_id 理论上不会，但防御性处理），取 mtime 最新的。
                """
                import glob as _glob

                pvc_name = os.environ.get("MODEL_PVC_NAME")
                host_models_path = os.environ.get("MODEL_HOST_PATH", "/models")
                prefix = f"sub_{submission_id}_"

                if not pvc_name:
                    # hostPath 模式：文件在 /models/uploads/
                    uploads_dir = os.path.join(host_models_path, "uploads")
                    candidates = _glob.glob(os.path.join(uploads_dir, f"{prefix}*"))
                    if not candidates:
                        return None
                    # 取 mtime 最新的（防御性，正常只有一个）
                    latest = max(candidates, key=os.path.getmtime)
                    basename = os.path.basename(latest)
                    return f"/models/uploads/{basename}"
                else:
                    # PVC 模式：文件在 UPLOADS_HOST_PATH，通过 /internal/uploads 暴露
                    candidates = _glob.glob(os.path.join(UPLOADS_HOST_PATH, f"{prefix}*"))
                    if not candidates:
                        return None
                    latest = max(candidates, key=os.path.getmtime)
                    basename = os.path.basename(latest)
                    api_base = os.environ.get("API_INTERNAL_URL", "http://leaderboard-api-svc:80")
                    return f"{api_base.rstrip('/')}/internal/uploads/{basename}"

            # 从首次提交时持久化的 algo_env_json 中恢复 qa_file_env，
            # 同时复用 SKIP_ALGO=1，确保 evaluator 跳过等待 algo /qa 的流程。
            saved_env = {}
            try:
                if original_submission.algo_env_json:
                    saved_env = json.loads(original_submission.algo_env_json)
            except Exception:
                saved_env = {}

            # 第一优先：从新版持久化的 algo_env_json 读
            qa_file_env = saved_env.get("USER_EVAL_FILE") or saved_env.get("QA_FILE")

            # 第二优先：存量兼容——按原始 sub.id 在磁盘上反查文件
            if not qa_file_env:
                qa_file_env = _recover_upload_qa_file(original_submission.id)
                if qa_file_env:
                    app.logger.info(
                        f"Rerun sub {sub_id}: algo_env_json missing, "
                        f"recovered qa_file_env from disk: {qa_file_env}"
                    )

            if not qa_file_env:
                db.session.rollback()
                return jsonify({"error": "Cannot rerun: upload file not found in algo_env_json or on disk"}), 400

            leaderboard = Leaderboard.query.get(new_submission.leaderboard_id)
            evaluator_image = (
                leaderboard.evaluator_image
                or os.environ.get("DEFAULT_EVALUATOR_IMAGE")
            )
            if not evaluator_image:
                db.session.rollback()
                return jsonify({"error": "evaluator_image not configured"}), 500

            try:
                resource_spec = json.loads(leaderboard.resource_spec) if leaderboard.resource_spec else {}
            except Exception:
                resource_spec = {}

            job_name = f"eval-sub-{new_submission.id}-{uuid.uuid4().hex[:6]}"
            new_submission.k8s_job_name = job_name

            # saved_env 包含 SKIP_ALGO=1 和 USER_EVAL_FILE（新版任务）；
            # 历史任务 saved_env 为空，但 SKIP_ALGO 由 create_job_manifest 内部在
            # start_algo_init=False 时自动注入，qa_file_env 由 qa_file_env 参数传入，不受影响。
            job_obj = create_job_manifest(
                job_name=job_name,
                leaderboard_id=new_submission.leaderboard_id,
                evaluator_image=evaluator_image,
                algorithm_image=None,
                resource_spec=resource_spec,
                submission_id=new_submission.id,
                extra_env=saved_env,      # 包含 SKIP_ALGO=1 和 USER_EVAL_FILE（新版有，历史版为空 dict 也无妨）
                start_algo_init=False,    # 关键：不创建 submitter-container，避免 ErrImageNeverPull
                qa_file_env=qa_file_env,  # 关键：QA_FILE 注入 evaluator，告知评测文件位置
            )

            if not k8s_batch_v1:
                db.session.rollback()
                return jsonify({"error": "Kubernetes client not initialized"}), 503

            k8s_batch_v1.create_namespaced_job(namespace=K8S_NAMESPACE, body=job_obj)
            new_submission.status = 'Pending'

        else:
            # ========== 镜像打榜 rerun ==========
            # Rerun 沿用原任务的 ENV 快照；注意不能用 `extra_env or None`，
            # 空 dict 时直接传 {} 而非 None，避免 ENV 被静默丢弃。
            extra_env = {}
            try:
                if original_submission.algo_env_json:
                    extra_env = json.loads(original_submission.algo_env_json)
            except Exception:
                extra_env = {}
            job_name = _start_k8s_job(new_submission, extra_env=extra_env)

        db.session.commit()
        return jsonify({
            "message": "Rerun requested. New evaluation job started.",
            "original_submission_id": sub_id,
            "new_submission_id": new_submission.id,
            "job_name": job_name,
            "status": new_submission.status
        }), 201
    except (ApiException, Exception) as e:
        db.session.rollback()
        error_detail = getattr(e, 'body', str(e))
        return jsonify({"error": "Failed to create K8s job for rerun", "details": error_detail}), 500


@app.route('/api/submission/<int:sub_id>', methods=['DELETE'])
@jwt_required()
def cancel_submission(sub_id):
    submission = Submission.query.get_or_404(sub_id)

    if not is_admin_or_owner(submission.user_id):
        return jsonify({"msg": "Forbidden: You are not the owner or an admin"}), 403

    current_status = _sync_submission_status(submission)
    if current_status not in ['Pending', 'Running', 'Submitted']:
        return jsonify({"msg": f"Cannot cancel a submission with status '{current_status}'"}), 400

    job_deleted = False
    if submission.k8s_job_name and k8s_batch_v1:
        try:
            # 取消前尽力抓一次日志（有就赚到）
            try:
                _persist_submission_logs(submission)
            except Exception:
                pass

            k8s_batch_v1.delete_namespaced_job(
                name=submission.k8s_job_name,
                namespace=K8S_NAMESPACE,
                body=client.V1DeleteOptions(propagation_policy='Background')
            )
            job_deleted = True
            app.logger.info(f"Deleted K8s job {submission.k8s_job_name} for cancelling submission {sub_id}")
        except ApiException as e:
            if e.status == 404:
                app.logger.warning(f"Job {submission.k8s_job_name} not found while cancelling sub {sub_id}, proceeding to mark DB.")
                job_deleted = True
            else:
                app.logger.error(f"Error deleting K8s job {submission.k8s_job_name} for sub {sub_id}: {e}")
                return jsonify({"msg": "Failed to delete underlying K8s job", "details": e.body}), 500
        except Exception as e:
            app.logger.error(f"Unexpected error deleting K8s job for sub {sub_id}: {e}")
            return jsonify({"msg": "Unexpected error deleting K8s job"}), 500

    submission.status = 'Cancelled'
    db.session.commit()

    # 取消后再试一次固化（大概率拿不到，但无伤大雅）
    try:
        _persist_submission_logs(submission)
    except Exception:
        pass

    return jsonify({"msg": f"Submission {sub_id} cancelled.", "k8s_job_deleted": job_deleted})


@app.route('/api/queue/status', methods=['GET'])
@jwt_required()
def get_queue_status():
    if not k8s_core_v1:
        return jsonify({"msg": "Kubernetes API client not available"}), 503
    try:
        pod_list = k8s_core_v1.list_namespaced_pod(
            namespace=K8S_NAMESPACE,
            label_selector="app=leaderboard-eval",
            field_selector="status.phase=Pending"
        )
        pending_count = len(pod_list.items)
        running_pod_list = k8s_core_v1.list_namespaced_pod(
            namespace=K8S_NAMESPACE,
            label_selector="app=leaderboard-eval",
            field_selector="status.phase=Running"
        )
        running_count = len(running_pod_list.items)
        return jsonify({"pending_tasks": pending_count, "running_tasks": running_count})
    except ApiException as e:
        app.logger.error(f"Error getting queue status: {e}")
        return jsonify({"msg": "Error getting K8s status", "details": e.body}), 500


# ======================================================================
# 模块 4: 公开榜单模块 (Public Display)
# ======================================================================

from sqlalchemy import func

@app.route('/api/leaderboards', methods=['GET'])
def get_public_leaderboards():
    # 1) 每个榜单的提交次数
    cnt_subq = (
        db.session.query(
            Submission.leaderboard_id.label("lb_id"),
            func.count(Submission.id).label("submission_count"),
        )
        .group_by(Submission.leaderboard_id)
        .subquery()
    )

    # 2) 每个榜单的“当前最佳分”（Succeeded 且 score 非空）
    best_subq = (
        db.session.query(
            Submission.leaderboard_id.label("lb_id"),
            func.max(Submission.score).label("current_sota_score"),
        )
        .filter(Submission.status == "Succeeded", Submission.score.isnot(None))
        .group_by(Submission.leaderboard_id)
        .subquery()
    )

    rows = (
        db.session.query(
            Leaderboard,
            func.coalesce(cnt_subq.c.submission_count, 0).label("submission_count"),
            best_subq.c.current_sota_score.label("current_sota_score"),
        )
        .outerjoin(cnt_subq, Leaderboard.id == cnt_subq.c.lb_id)
        .outerjoin(best_subq, Leaderboard.id == best_subq.c.lb_id)
        .order_by(Leaderboard.name)
        .all()
    )

    board_list = []
    for b, submission_count, current_sota_score in rows:
        board_list.append(
            {
                "id": b.id,
                "name": b.name,
                "description": b.description,
                "version": b.version,
                "owner_username": b.owner.username,
                "difficulty_factor": b.difficulty_factor,

                # 你原来的字段：配置/基准 SOTA（不自动变，除非你手动或代码更新）
                "sota_score": b.sota_score,

                # 新增字段：实时计算的“当前榜单最佳分”（用于前端 SOTA 标签展示）
                "current_sota_score": current_sota_score,

                # 新增字段：提交次数
                "submission_count": int(submission_count or 0),
            }
        )

    return jsonify(board_list)


@app.route('/api/leaderboard/<int:leaderboard_id>/rankings', methods=['GET'])
def get_public_rankings(leaderboard_id):
    leaderboard = Leaderboard.query.get(leaderboard_id)
    if not leaderboard:
        return jsonify({"msg": f"Leaderboard {leaderboard_id} not found"}), 404

    # 轻量同步：把该榜单最近的非终态任务拉一遍 K8s 状态
    non_terminal = (
        db.session.query(Submission)
        .filter(
            Submission.leaderboard_id == leaderboard_id,
            Submission.status.in_(('Submitted', 'Pending', 'Running'))
        )
        .order_by(Submission.submitted_at.desc())
        .limit(50)
        .all()
    )
    for sub in non_terminal:
        _sync_submission_status(sub)

    per_submission = request.args.get('per_submission') in ('1', 'true', 'True')
    include_job = request.args.get('include_job') in ('1', 'true', 'True')
    try:
        limit = int(request.args.get('limit')) if request.args.get('limit') else None
    except ValueError:
        limit = None

    if per_submission:
        q = (
            db.session.query(
                Submission.id.label('submission_id'),
                Submission.submission_name,
                Submission.k8s_job_name,
                Submission.algorithm_image_url,
                Submission.score,
                Submission.metrics_json,
                Submission.submitted_at,
                Submission.algo_env_json,
                Submission.algo_env_grid_keys,
                User.id.label('user_id'),
                User.username
            )
            .join(User, User.id == Submission.user_id)
            .filter(
                Submission.leaderboard_id == leaderboard_id,
                Submission.status == 'Succeeded',
                Submission.score.isnot(None)
            )
            .order_by(db.desc(Submission.score), db.asc(Submission.submitted_at))
        )
        if limit:
            q = q.limit(limit)

        rows = q.all()
        ranking_list = []
        for idx, r in enumerate(rows, start=1):
            metrics = None
            if r.metrics_json:
                try:
                    metrics = json.loads(r.metrics_json)
                except Exception:
                    metrics = None

            algo_env = None
            algo_env_grid_keys = None
            if r.algo_env_json:
                try:
                    algo_env = json.loads(r.algo_env_json)
                except Exception:
                    algo_env = None
            if r.algo_env_grid_keys:
                try:
                    algo_env_grid_keys = json.loads(r.algo_env_grid_keys)
                except Exception:
                    algo_env_grid_keys = None

            item = {
                "rank": idx,
                "user_id": r.user_id,
                "username": r.username,
                "score": r.score,
                "metrics": metrics,
                "last_submitted": r.submitted_at.isoformat(),
            }
            if include_job:
                item["job_name"] = r.k8s_job_name
                item["algorithm_image_url"] = r.algorithm_image_url
                item["submission_name"] = r.submission_name
                item["submission_id"] = r.submission_id
            if algo_env is not None:
                item["algo_env"] = algo_env
            if algo_env_grid_keys is not None:
                item["algo_env_grid_keys"] = algo_env_grid_keys

            ranking_list.append(item)

        return jsonify({"leaderboard_name": leaderboard.name, "rankings": ranking_list})

    else:
        rankings = (
            db.session.query(
                Submission.user_id,
                User.username,
                db.func.max(Submission.score).label('best_score'),
                db.func.min(Submission.submitted_at).label('first_success_time'),
                db.func.max(Submission.submitted_at).label('last_submission_time')
            )
            .join(User, User.id == Submission.user_id)
            .filter(
                Submission.leaderboard_id == leaderboard_id,
                Submission.status == 'Succeeded',
                Submission.score.isnot(None)
            )
            .group_by(Submission.user_id, User.username)
            .order_by(db.desc('best_score'), db.asc('first_success_time'))
            .all()
        )

        ranking_list = []
        for i, r in enumerate(rankings):
            item = {
                "rank": i + 1,
                "user_id": r.user_id,
                "username": r.username,
                "score": r.best_score,
                "last_submitted": r.last_submission_time.isoformat() if r.last_submission_time else None
            }
            if include_job:
                best_sub = (
                    Submission.query
                    .filter(
                        Submission.leaderboard_id == leaderboard_id,
                        Submission.user_id == r.user_id,
                        Submission.status == 'Succeeded',
                        Submission.score == r.best_score
                    )
                    .order_by(Submission.submitted_at.asc())
                    .first()
                )
                if best_sub:
                    item["job_name"] = best_sub.k8s_job_name
                    item["algorithm_image_url"] = best_sub.algorithm_image_url
                    item["submission_name"] = best_sub.submission_name
                    item["submission_id"] = best_sub.id
            ranking_list.append(item)

        return jsonify({"leaderboard_name": leaderboard.name, "rankings": ranking_list})


# ====================== 榜单统计（平均用时） ======================
@app.route('/api/leaderboard/<int:leaderboard_id>/stats', methods=['GET'])
def leaderboard_stats(leaderboard_id):
    """
    返回榜单平均用时（秒）。
    优先：SubmissionLog.finalized_at - Submission.submitted_at；
    回退：metrics_json 里的 Avg Time 字段均值；
    否则：avg_duration_sec=None, method='none'
    """
    board = Leaderboard.query.get(leaderboard_id)
    if not board:
        return jsonify({"msg": f"Leaderboard {leaderboard_id} not found"}), 404

    # 1) 日志法
    q = (
        db.session.query(Submission, SubmissionLog)
        .join(SubmissionLog, SubmissionLog.submission_id == Submission.id)
        .filter(
            Submission.leaderboard_id == leaderboard_id,
            Submission.status == 'Succeeded',
            SubmissionLog.finalized_at.isnot(None)
        )
        .order_by(Submission.submitted_at.desc())
        .limit(500)
    )
    rows = q.all()
    durations = []
    for sub, slog in rows:
        if sub.submitted_at and slog.finalized_at:
            dur = (slog.finalized_at - sub.submitted_at).total_seconds()
            if 0 <= dur <= 7 * 24 * 3600:  # 合理范围
                durations.append(dur)
    if durations:
        avg_log = sum(durations) / len(durations)
        return jsonify({
            "leaderboard_id": leaderboard_id,
            "avg_duration_sec": float(avg_log),
            "sample_count": len(durations),
            "method": "log"
        })

    # 2) 指标法（从 metrics_json 猜测）
    alt_keys = ["Avg Time (s)", "avg_time_s", "avg_time", "duration_s"]
    q2 = (
        db.session.query(Submission)
        .filter(
            Submission.leaderboard_id == leaderboard_id,
            Submission.status == 'Succeeded',
            Submission.metrics_json.isnot(None)
        )
        .order_by(Submission.submitted_at.desc())
        .limit(500)
    )
    vals = []
    for sub in q2.all():
        try:
            mj = json.loads(sub.metrics_json or "{}")
            if isinstance(mj, dict):
                for k in alt_keys:
                    if k in mj and mj[k] is not None:
                        vals.append(float(mj[k]))
                        break
        except Exception:
            continue
    if vals:
        avg_m = sum(vals) / len(vals)
        return jsonify({
            "leaderboard_id": leaderboard_id,
            "avg_duration_sec": float(avg_m),
            "sample_count": len(vals),
            "method": "metrics"
        })

    # 3) 无法估计
    return jsonify({
        "leaderboard_id": leaderboard_id,
        "avg_duration_sec": None,
        "sample_count": 0,
        "method": "none"
    })

from flask import abort

def _current_user_from_jwt():
    ident = get_jwt_identity()
    if ident is None:
        return None
    try:
        if isinstance(ident, int) or (isinstance(ident, str) and ident.isdigit()):
            return User.query.get(int(ident))
    except Exception:
        pass
    return User.query.filter_by(username=str(ident)).first()

def _require_admin():
    u = _current_user_from_jwt()
    if not u or u.role != "admin":
        abort(403, description="Admin privilege required")
    return u

@app.route("/api/admin/submission/<int:submission_id>", methods=["DELETE"])
@jwt_required()
def admin_delete_submission(submission_id: int):
    admin_user = _require_admin()

    sub = Submission.query.get(submission_id)
    if not sub:
        return jsonify({"ok": False, "msg": f"submission {submission_id} not found"}), 404

    # best-effort 删除 K8s Job（有就删；没有就算了）
    if sub.k8s_job_name and k8s_batch_v1:
        try:
            k8s_batch_v1.delete_namespaced_job(
                name=sub.k8s_job_name,
                namespace=K8S_NAMESPACE,
                propagation_policy="Background",
                grace_period_seconds=0,
            )
        except Exception:
            pass

    # 回收积分事件（best-effort 回扣月度分）
    try:
        evs = PointsEvent.query.filter_by(submission_id=sub.id).all()
        for ev in evs:
            mp = MonthlyPoints.query.filter_by(user_id=ev.user_id, year_month=ev.year_month).first()
            if mp:
                try:
                    mp.total_points = float(mp.total_points or 0.0) - float(ev.points or 0.0)
                    if mp.total_points < 0:
                        mp.total_points = 0.0
                except Exception:
                    pass
            db.session.delete(ev)
    except Exception:
        pass

    # 清理明细与日志
    try:
        SubmissionEvalDetail.query.filter_by(submission_id=sub.id).delete(synchronize_session=False)
    except Exception:
        pass
    try:
        SubmissionLog.query.filter_by(submission_id=sub.id).delete(synchronize_session=False)
    except Exception:
        pass

    lbid = sub.leaderboard_id
    try:
        db.session.delete(sub)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "msg": f"delete failed: {e}"}), 500

    return jsonify({
        "ok": True,
        "deleted_submission_id": submission_id,
        "leaderboard_id": lbid,
        "by_admin": admin_user.username,
    })



# ======================================================================
# K8s Job 创建辅助函数（sidecar 方式）
# ======================================================================
def create_job_manifest(
    job_name,
    leaderboard_id,
    evaluator_image,
    algorithm_image,
    resource_spec,
    submission_id,
    extra_env: dict | None = None,
    start_algo_init: bool = True,   # 新增：是否需要作为 init container 启动算法容器（默认 True）
    qa_file_env: str | None = None  # 新增：传入给 evaluator 的 QA_FILE（本地路径或 http url）
):
    # === /models 卷定义（默认 hostPath，可选 PVC，通过环境变量控制）===
    pvc_name = os.environ.get("MODEL_PVC_NAME")          # 若提供则使用 PVC
    host_models_path = os.environ.get("MODEL_HOST_PATH", "/models")  # 默认宿主机路径

    if pvc_name:
        models_volume = client.V1Volume(
            name="models-vol",
            persistent_volume_claim=client.V1PersistentVolumeClaimVolumeSource(
                claim_name=pvc_name, read_only=False
            ),
        )
    else:
        models_volume = client.V1Volume(
            name="models-vol",
            host_path=client.V1HostPathVolumeSource(path=host_models_path, type="Directory"),
        )
    shared_volume = client.V1Volume(
        name="shared-volume",
        empty_dir=client.V1EmptyDirVolumeSource()
    )

    models_mount = client.V1VolumeMount(name="models-vol", mount_path="/models")
    shared_mount = client.V1VolumeMount(name="shared-volume", mount_path="/shared")

    # 通用产物目录（不耦合 phase）：/models/xarena/<bench_slug>/<jobid>
    # 两容器都挂了 /models（hostPath，跨 pod 持久）；各自把输出拷到 $OUTPUT_DIR/{eval,algo}。
    _board = Leaderboard.query.get(leaderboard_id)
    _bench_slug = re.sub(r"[^a-z0-9]+", "-", ((_board.name if _board else "") or "").lower()).strip("-")[:40] or f"lb{leaderboard_id}"
    output_dir = f"/models/xarena/{_bench_slug}/{job_name}"

    # 准备注入的额外 ENV（已做优先级合并与过滤）
    extra_env = _sanitize_env_dict(extra_env or {})
    extra_env_vars = _dict_to_envvars(extra_env)

    # 基础 ENV（保留关键，不允许被覆盖）
    base_eval_env = [
        client.V1EnvVar(name="SUBMISSION_ID", value=str(submission_id)),
        client.V1EnvVar(name="LEADERBOARD_ID", value=str(leaderboard_id)),
        client.V1EnvVar(name="API_INTERNAL_URL", value=os.environ.get("API_INTERNAL_URL", "http://leaderboard-api-svc:80")),
        client.V1EnvVar(name="ALGO_API_ENDPOINT", value=os.environ.get("ALGO_API_ENDPOINT", "http://localhost:5000")),
        client.V1EnvVar(name="OUTPUT_DIR", value=output_dir),
    ]
    # 如果给定 qa_file_env，注入 QA_FILE（覆盖容器默认），这是 evaluator 将读取的入口
    if qa_file_env:
        base_eval_env.append(client.V1EnvVar(name="QA_FILE", value=str(qa_file_env)))

    # base algo env （仅当需要启动算法容器时才有意义）
    base_algo_env = [
        client.V1EnvVar(name="SUBMISSION_ID", value=str(submission_id)),
        client.V1EnvVar(name="LEADERBOARD_ID", value=str(leaderboard_id)),
        client.V1EnvVar(name="API_INTERNAL_URL", value=os.environ.get("API_INTERNAL_URL", "http://leaderboard-api-svc:80")),
        client.V1EnvVar(name="OUTPUT_DIR", value=output_dir),
    ]

    # EVAL container
    # 当 start_algo_init=False 时，注入 SKIP_ALGO=1（让 evaluator 跳过等待 algo /qa 的逻辑）
    eval_extra_env = []
    if not start_algo_init:
        eval_extra_env.append(client.V1EnvVar(name="SKIP_ALGO", value="1"))

    evaluator_container = client.V1Container(
        name="evaluator-container",
        image=evaluator_image,
        image_pull_policy=os.environ.get("K8S_IMAGE_PULL_POLICY", "Never"),
        resources=client.V1ResourceRequirements(
            requests={"cpu": os.environ.get("K8S_EVAL_CPU_REQ", "1"),
                      "memory": os.environ.get("K8S_EVAL_MEM_REQ", "2Gi")},
            limits={"cpu": os.environ.get("K8S_EVAL_CPU_LIMIT", "3"),
                    "memory": os.environ.get("K8S_EVAL_MEM_LIMIT", "6Gi")},
        ),
        env=base_eval_env + extra_env_vars + eval_extra_env,
        volume_mounts=[models_mount, shared_mount],
    )

    # ALGO (init sidecar) - 仅在 start_algo_init=True 时创建
    algorithm_container = None
    if start_algo_init:
        algorithm_container = client.V1Container(
            name="submitter-container",
            image=algorithm_image,
            image_pull_policy=os.environ.get("K8S_IMAGE_PULL_POLICY", "Never"),
            resources=client.V1ResourceRequirements(
                limits=resource_spec.get("limits"),
                requests=resource_spec.get("requests")
            ),
            restart_policy="Always",
            env=base_algo_env + extra_env_vars,
            volume_mounts=[models_mount, shared_mount],
        )

    # imagePullSecrets (如果需要从私有仓库拉取)
    pull_secrets = []
    secret_name = os.environ.get("K8S_IMAGE_PULL_SECRET")
    if secret_name:
        pull_secrets.append(client.V1LocalObjectReference(name=secret_name))

    # 注入 job pod 的密钥（deepseek/dashscope）：当 K8S_JOB_SECRET 设置时给两容器加 envFrom secretRef
    _job_secret = os.environ.get("K8S_JOB_SECRET")
    if _job_secret:
        _ef = [client.V1EnvFromSource(secret_ref=client.V1SecretEnvSource(name=_job_secret, optional=True))]
        evaluator_container.env_from = _ef
        if algorithm_container is not None:
            algorithm_container.env_from = _ef

    # PodSpec：如果 start_algo_init=True 则把算法容器作为 init_containers
    pod_spec_kwargs = {
        "containers": [evaluator_container],
        "restart_policy": "Never",
        "service_account_name": K8S_SERVICE_ACCOUNT,
        "priority_class_name": os.environ.get("K8S_PRIORITY_CLASS", "leaderboard-job-priority"),
        "image_pull_secrets": pull_secrets if pull_secrets else None,
        "volumes": [models_volume, shared_volume],
    }
    if start_algo_init and algorithm_container:
        pod_spec_kwargs["init_containers"] = [algorithm_container]
    else:
        # 不需要算法 init container 时，明确不设置 init_containers（默认为 None）
        pod_spec_kwargs["init_containers"] = None

    pod_template_spec = client.V1PodTemplateSpec(
        metadata=client.V1ObjectMeta(labels={"app": "leaderboard-eval", "job-name": job_name}),
        spec=client.V1PodSpec(**pod_spec_kwargs)
    )

    # backoff_limit=2：至多重试 0 次（共 1 次），防止无限重试
    job_spec = client.V1JobSpec(
        template=pod_template_spec,
        backoff_limit=0,
        active_deadline_seconds=K8S_JOB_ACTIVE_DEADLINE,
        ttl_seconds_after_finished=K8S_JOB_TTL_SECONDS
    )

    job = client.V1Job(
        api_version="batch/v1",
        kind="Job",
        metadata=client.V1ObjectMeta(name=job_name, namespace=K8S_NAMESPACE),
        spec=job_spec
    )
    return job


@app.route("/internal/uploads/<path:filename>", methods=["GET"])
def _serve_internal_uploads(filename):
    # 注意：该接口应只在内部网络或认证下使用
    return send_from_directory(UPLOADS_HOST_PATH, filename, as_attachment=True)


@app.route('/api/submission/<int:sub_id>/extra', methods=['GET'])
@jwt_required()
def get_submission_extra(sub_id):
    """查看某次提交回传的 extraconfig（submission_eval_details.extra_json）。"""
    sub = Submission.query.get(sub_id)
    if not sub:
        return jsonify({"msg": "submission not found"}), 404
    uid = get_jwt_identity()
    me = User.query.get(uid)
    board = Leaderboard.query.get(sub.leaderboard_id)
    is_admin = bool(me and me.role == 'admin')
    is_owner = (str(sub.user_id) == str(uid))
    is_board_owner = bool(board and str(board.owner_id) == str(uid))
    if not (is_admin or is_owner or is_board_owner):
        return jsonify({"msg": "forbidden"}), 403
    qid = request.args.get('question_id')
    q = SubmissionEvalDetail.query.filter_by(submission_id=sub_id)
    if qid:
        q = q.filter_by(question_id=str(qid))
    rows = q.limit(int(request.args.get('limit', 200))).all()
    items = []
    for r in rows:
        try:
            extra = json.loads(r.extra_json) if r.extra_json else None
        except Exception:
            extra = None
        items.append({"question_id": r.question_id, "is_correct": r.is_correct,
                      "pred_answer": r.pred_answer, "extra": extra})
    return jsonify({"submission_id": sub_id, "count": len(items), "items": items})

@app.route("/api/submission/upload", methods=["POST"])
@jwt_required(optional=True)
def upload_submission_and_submit_job():
    """
    上传现成的评测 JSON 文件并触发 Job（跳过 algo sidecar）。

    支持：
      - 前端：带 JWT，自动用当前登录用户
      - curl：可以不带 JWT，只要传 user_id，或配置 UPLOAD_DEFAULT_USER_ID

    表单字段：
      - leaderboard_id (required)
      - title (optional) -> 映射到 submission_name（同一 user 下需唯一）
      - file (required, multipart/form-data)
      - resource_spec (optional, json string)
      - evaluator_image (optional，默认用对应榜单的 evaluator_image)
      - user_id (optional，curl 时可显式指定；优先级最高)
    """
    import time
    # ========== 1) 解析榜单 ID ==========
    lb_raw = request.form.get("leaderboard_id")
    if not lb_raw and request.is_json:
        lb_raw = (request.json or {}).get("leaderboard_id")
    if not lb_raw:
        return jsonify({"msg": "missing leaderboard_id"}), 400

    try:
        lb_id = int(lb_raw)
    except Exception:
        return jsonify({"msg": "invalid leaderboard_id"}), 400

    board = Leaderboard.query.get(lb_id)
    if not board:
        return jsonify({"msg": f"Leaderboard {lb_id} not found"}), 404

    # ========== 2) 决定 user_id（不再 401） ==========
    user_id = None

    # 2.1 curl 可以直接传 user_id
    raw_uid = request.form.get("user_id")
    if raw_uid:
        try:
            user_id = int(raw_uid)
        except Exception:
            user_id = None

    # 2.2 前端带 JWT 的情况：用当前登录用户
    if user_id is None:
        ident = get_jwt_identity()
        if ident is not None:
            try:
                user_id = int(ident)
            except Exception:
                user_id = None

    # 2.3 兜底：环境变量 UPLOAD_DEFAULT_USER_ID（可在部署时设）
    if user_id is None:
        env_uid = os.environ.get("UPLOAD_DEFAULT_USER_ID")
        if env_uid:
            try:
                user_id = int(env_uid)
            except Exception:
                user_id = None

    if user_id is None:
        # 不返回 401，避免前端误判为登录过期
        return jsonify({
            "msg": "cannot determine user_id: 请登录，或表单中传 user_id，或配置环境变量 UPLOAD_DEFAULT_USER_ID"
        }), 400

    # ========== 3) 取文件 ==========
    if "file" not in request.files:
        return jsonify({"msg": "no file uploaded (use multipart/form-data with 'file')"}), 400

    f = request.files["file"]
    if not f or not f.filename:
        return jsonify({"msg": "empty file"}), 400

    filename = secure_filename(f.filename) or f"upload_{int(time.time())}.json"

    # ========== 4) submission_name & 唯一性校验 ==========
    title = request.form.get("title") or f"upload-eval-{int(time.time())}"
    existing = Submission.query.filter_by(user_id=user_id, submission_name=title).first()
    if existing:
        return jsonify({
            "msg": "Submission name already used by this user.",
            "submission_name": title
        }), 409

    # 先建 Submission 记录，拿到 ID 用来命名文件
    sub = Submission(
        leaderboard_id=lb_id,
        submission_name=title,
        algorithm_image_url="upload-mode",  # 这里只是占位，不会真的拉 algo 镜像
        status="Created",
        user_id=user_id,
    )
    db.session.add(sub)
    db.session.flush()  # 拿到 sub.id，但还不 commit，后面出错可以整体回滚

    # ========== 5) 落地文件 ==========
    pvc_name = os.environ.get("MODEL_PVC_NAME")
    host_models_path = os.environ.get("MODEL_HOST_PATH", "/models")
    qa_file_env = None

    if not pvc_name:
        # hostPath 情况：直接写 /models/uploads，评测容器也挂载这个目录
        uploads_dir = os.path.join(host_models_path, "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        dest_basename = f"sub_{sub.id}_{int(time.time())}_{filename}"
        dest_full = os.path.join(uploads_dir, dest_basename)
        f.save(dest_full)
        qa_file_env = f"/models/uploads/{dest_basename}"
    else:
        # PVC 情况：写到 UPLOADS_HOST_PATH，再通过 /internal/uploads 暴露
        dest_basename = f"sub_{sub.id}_{int(time.time())}_{filename}"
        os.makedirs(UPLOADS_HOST_PATH, exist_ok=True)
        dest_full = os.path.join(UPLOADS_HOST_PATH, dest_basename)
        f.save(dest_full)
        api_base = os.environ.get("API_INTERNAL_URL", "http://leaderboard-api-svc:80")
        qa_file_env = f"{api_base.rstrip('/')}/internal/uploads/{dest_basename}"

    # 没有对应列也没关系，只是挂个属性，方便你调试
    try:
        sub.dataset_path = dest_full  # type: ignore[attr-defined]
    except Exception:
        pass

    # ========== 5.5) 持久化 upload-mode 的关键参数，供 rerun 使用 ==========
    # rerun 时需要知道：评测文件路径（qa_file_env）、以及该走 start_algo_init=False 路径。
    # algorithm_image_url="upload-mode" 已经标记了提交类型，此处只需把文件路径存进去。
    # SKIP_ALGO 和 USER_EVAL_FILE 与首次提交调用 create_job_manifest 时的 extra_env 保持一致，
    # 这样 rerun 可以原样复用这份快照重建 Job。
    sub.algo_env_json = json.dumps(
        {"SKIP_ALGO": "1", "USER_EVAL_FILE": qa_file_env},
        ensure_ascii=False
    )

    # ========== 6) 选 evaluator 镜像 & 资源规格 ==========
    evaluator_image = (
        request.form.get("evaluator_image")
        or os.environ.get("EVALUATOR_IMAGE")
        or board.evaluator_image
        or os.environ.get("DEFAULT_EVALUATOR_IMAGE")
    )
    if not evaluator_image:
        db.session.rollback()
        return jsonify({"msg": "evaluator_image not configured (表单 / 环境变量 / 榜单 都为空)"}), 500

    # resource_spec：表单优先，其次用榜单上的配置，最后给个小默认
    try:
        rs_raw = request.form.get("resource_spec")
        if rs_raw:
            resource_spec = json.loads(rs_raw)
        else:
            resource_spec = json.loads(board.resource_spec) if board.resource_spec else {}
    except Exception:
        resource_spec = {}

    if not resource_spec:
        resource_spec = {
            "limits": {"cpu": "1", "memory": "2Gi"},
            "requests": {"cpu": "500m", "memory": "1Gi"},
        }

    # ========== 7) 构造并提交 K8s Job（只跑 evaluator-container） ==========
    job_name = f"eval-upload-{sub.id}-{int(time.time())}"
    job_obj = create_job_manifest(
        job_name=job_name,
        leaderboard_id=lb_id,
        evaluator_image=evaluator_image,
        algorithm_image=None,
        resource_spec=resource_spec,
        submission_id=sub.id,
        extra_env={"SKIP_ALGO": "1", "USER_EVAL_FILE":qa_file_env},
        start_algo_init=False,   # 关键：不启动 algo init 容器
        qa_file_env=qa_file_env, # 关键：QA_FILE 传给 evaluator
    )

    if not k8s_batch_v1:
        db.session.rollback()
        return jsonify({"msg": "Kubernetes client not initialized"}), 503

    try:
        k8s_batch_v1.create_namespaced_job(namespace=K8S_NAMESPACE, body=job_obj)
    except Exception as e:
        sub.status = "Failed"
        db.session.commit()
        return jsonify({"msg": f"Failed to create k8s job: {e}"}), 500

    sub.k8s_job_name = job_name
    sub.status = "Running"
    db.session.commit()

    return jsonify({"submission_id": sub.id, "job_name": job_name, "qa_file": qa_file_env}), 200


# ======================================================================
# 内部 API：评测过程明细记录 + 回写分数/指标 + 触发积分统计
# ======================================================================
def _year_month(dt: datetime) -> str:
    return dt.strftime("%Y-%m")


def _compute_points(leaderboard: Leaderboard, score: float):
    """
    分段计分：
    - 仅当 score > SOTA 时才有积分；
    - [SOTA, 80) 区间的超越部分：1x；
    - [80, +∞) 区间的超越部分：2x；
    - 最终 points = (low_part * 1 + high_part * 2) * difficulty_factor
    返回: (points, delta_total, effective_multiplier)
      - delta_total = max(0, score - SOTA)
      - effective_multiplier 若存在高区间贡献则记 2.0，否则 1.0（为兼容 PointsEvent 里历史字段）
    """
    if leaderboard is None or leaderboard.sota_score is None:
        return 0.0, 0.0, 1.0

    try:
        s = float(score)
        s_sota = float(leaderboard.sota_score)
    except Exception:
        return 0.0, 0.0, 1.0

    # 未超过 SOTA 不计分
    if s <= s_sota:
        return 0.0, 0.0, 1.0

    # 分段：低区间 [SOTA, 80) -> 1x；高区间 [max(80, SOTA), s] -> 2x
    low_part = 0.0
    high_part = 0.0

    # 低区间上界
    low_upper = min(s, 80.0)
    if low_upper > s_sota:
        low_part = low_upper - s_sota  # 只要 SOTA < 80，且分数超过 SOTA，就会有低区间贡献

    # 高区间从 max(80, SOTA) 开始
    if s > 80.0:
        high_start = max(80.0, s_sota)
        if s > high_start:
            high_part = s - high_start

    difficulty = float(leaderboard.difficulty_factor or 1.0)
    points = (low_part * 1.0 + high_part * 2.0) * difficulty
    delta_total = (s - s_sota)

    # 若有高区间贡献，标记 multiplier=2.0；否则 1.0（字段保留用于事件表展示）
    effective_multiplier = 2.0 if high_part > 0 else 1.0
    return points, delta_total, effective_multiplier


def _create_points_event_if_eligible(submission: Submission):
    """在提交成功且超过 SOTA 时产出积分事件并累加到月度汇总。"""
    if submission is None or submission.status != 'Succeeded' or submission.score is None:
        return
    leaderboard = submission.leaderboard
    if not leaderboard:
        return

    points, delta, multiplier = _compute_points(leaderboard, submission.score)
    if points <= 0:
        return

    ym = _year_month(datetime.utcnow())
    evt = PointsEvent(
        user_id=submission.user_id,
        leaderboard_id=submission.leaderboard_id,
        submission_id=submission.id,
        year_month=ym,
        points=points,
        delta=delta,
        multiplier=multiplier,
        difficulty_factor=float(leaderboard.difficulty_factor or 1.0),
        score=float(submission.score),
        sota_score=leaderboard.sota_score
    )
    db.session.add(evt)

    mp = MonthlyPoints.query.filter_by(user_id=submission.user_id, year_month=ym).first()
    if not mp:
        mp = MonthlyPoints(user_id=submission.user_id, year_month=ym, total_points=0.0)
        db.session.add(mp)
    mp.total_points = float(mp.total_points or 0.0) + float(points)

    db.session.commit()
    app.logger.info(f"[Points] user={submission.user_id} +{points:.4f} in {ym} (delta={delta:.4f}, mult={multiplier})")


@app.route('/api/internal/submission/<int:sub_id>/eval-detail', methods=['POST'])
def append_submission_eval_detail(sub_id):
    """
    评测脚本逐题回传明细：
    Body 大致形如（字段有就用，没有就忽略）：
    {
      "question_id": "...",   // 或 qid
      "question": "...",      // 或 q
      "gold_answer": "...",   // 或 gold / reference
      "pred_answer": "...",   // 或 answer / prediction / model_answer
      "is_correct": 1,        // 或 correct / result, 支持 0/1/bool/"0"/"1"
      "latency_ms": 123.4,
      "used_tokens": 100,
      "retrieved": 8,
      "eval_prompt": "...",   // 或 prompt
      "extra": {...}          // 任意 KV，会与原始 body 一起写入 extra_json
    }
    """
    submission = Submission.query.get(sub_id)
    if not submission:
        return jsonify({"msg": "Submission not found"}), 404

    data = request.get_json(force=True, silent=True) or {}

    qid = data.get("question_id") or data.get("qid")
    if qid is not None:
        qid = str(qid)

    question = data.get("question") or data.get("q") or ""
    gold_answer = (
        data.get("gold_answer")
        or data.get("gold")
        or data.get("reference")
        or data.get("std_answer")
        or ""
    )
    pred_answer = (
        data.get("pred_answer")
        or data.get("prediction")
        or data.get("answer")
        or data.get("model_answer")
        or ""
    )

    is_correct_raw = data.get("is_correct", data.get("correct", data.get("result")))
    is_correct = None
    if is_correct_raw is not None:
        if isinstance(is_correct_raw, bool):
            is_correct = is_correct_raw
        elif isinstance(is_correct_raw, (int, float)):
            is_correct = bool(int(is_correct_raw))
        elif isinstance(is_correct_raw, str):
            s = is_correct_raw.strip().lower()
            if s in ("1", "true", "yes", "y", "正确", "对"):
                is_correct = True
            elif s in ("0", "false", "no", "n", "错误", "错"):
                is_correct = False

    def _to_float(x):
        try:
            if x is None:
                return None
            return float(x)
        except Exception:
            return None

    def _to_int(x):
        try:
            if x is None:
                return None
            return int(x)
        except Exception:
            return None

    latency_ms = _to_float(data.get("latency_ms"))
    used_tokens = _to_int(data.get("used_tokens"))
    retrieved = _to_int(data.get("retrieved"))

    eval_prompt = data.get("eval_prompt") or data.get("prompt") or None

    # extra_json：保留原始 body，方便后续扩展/展开
    try:
        extra_json = json.dumps(data, ensure_ascii=False)
    except Exception:
        extra_json = None

    rec = SubmissionEvalDetail(
        submission_id=submission.id,
        leaderboard_id=submission.leaderboard_id,
        question_id=qid,
        question=question,
        gold_answer=gold_answer,
        pred_answer=pred_answer,
        is_correct=is_correct,
        latency_ms=latency_ms,
        used_tokens=used_tokens,
        retrieved=retrieved,
        eval_prompt=eval_prompt,
        extra_json=extra_json
    )
    db.session.add(rec)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error inserting eval-detail for sub {sub_id}: {e}")
        return jsonify({"msg": "DB error inserting eval detail"}), 500

    return jsonify({"msg": "Eval detail appended"})

@app.route('/api/internal/submission/<int:sub_id>/score', methods=['POST'])
def update_submission_score(sub_id):
    """evaluator 回调：接受 {score} 或 {metrics:{...,score:...}}，可选 status（不传则默认记为 Succeeded）"""
    submission = Submission.query.get(sub_id)
    if not submission:
        return jsonify({"msg": "Submission not found"}), 404

    payload = request.get_json(force=True, silent=True) or {}
    score_in_body = payload.get('score', None)
    metrics = payload.get('metrics', None)
    status_opt = payload.get('status', None)

    try:
        # 1) 兼容两种上报方式（原逻辑不动）
        if metrics is not None:
            if not isinstance(metrics, dict):
                return jsonify({"msg": "'metrics' must be a JSON object"}), 400
            submission.metrics_json = json.dumps(metrics, ensure_ascii=False)
            if 'score' in metrics and metrics['score'] is not None:
                submission.score = float(metrics['score'])
        elif score_in_body is not None:
            submission.score = float(score_in_body)
        else:
            return jsonify({"msg": "Missing 'score' or 'metrics' with 'score'"}), 400

        # 2) 可选：同时更新状态（原逻辑不动）
        if status_opt in ('Succeeded', 'Failed', 'Running', 'Pending', 'Submitted', 'Cancelled'):
            submission.status = status_opt
        else:
            submission.status = 'Succeeded'

        # 3) 极小增量：顺手把 eval_details 落到 submission_eval_details
        # ------------------------------------------------------
        # evaluator 现在会：
        #   payload["eval_details"] = [...]
        #   metrics["eval_details"] = [...]
        # 我们两个都兼容一下
        eval_details = payload.get("eval_details")
        if eval_details is None and isinstance(metrics, dict):
            eval_details = metrics.get("eval_details")

        if eval_details and isinstance(eval_details, list) and submission.leaderboard_id is not None:
            # 先删掉旧的明细，避免重复
            SubmissionEvalDetail.query.filter_by(submission_id=submission.id).delete()

            for d in eval_details:
                if not isinstance(d, dict):
                    continue

                extra = d.get("extra")
                extra_json = None
                if extra is not None:
                    try:
                        extra_json = json.dumps(extra, ensure_ascii=False)
                    except Exception:
                        extra_json = None

                ic_raw = d.get("is_correct", None)
                is_correct = None if ic_raw is None else bool(ic_raw)

                det = SubmissionEvalDetail(
                    submission_id=submission.id,
                    # ★ 关键：不要从 payload 取，用 submission 自己的 leaderboard_id
                    leaderboard_id=submission.leaderboard_id,
                    question_id=d.get("question_id"),
                    question=d.get("question"),
                    gold_answer=d.get("gold_answer"),
                    pred_answer=d.get("pred_answer"),
                    is_correct=is_correct,
                    latency_ms=float(d["latency_ms"]) if d.get("latency_ms") is not None else None,
                    used_tokens=int(d.get("used_tokens") or 0),
                    retrieved=int(d.get("retrieved") or 0),
                    eval_prompt=d.get("eval_prompt"),
                    extra_json=extra_json,
                )
                db.session.add(det)
        # ------------------------------------------------------

        db.session.commit()

        # 4) 状态为 Succeeded 时尝试计分 + 尝试固化日志（原逻辑不动）
        if submission.status == 'Succeeded':
            _create_points_event_if_eligible(submission)
            try:
                _persist_submission_logs(submission)
            except Exception as e:
                app.logger.error(f"Persist logs after score callback failed for sub {submission.id}: {e}")

        app.logger.info(f"Score/metrics updated for submission {sub_id}")
        return jsonify({"msg": "Score/metrics updated"})

    except ValueError:
        db.session.rollback()
        return jsonify({"msg": "'score' must be a valid number"}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating score for sub {sub_id}: {e}")
        return jsonify({"msg": "Database error updating score"}), 500

# ======================================================================
# 模块 5: 积分查询接口
# ======================================================================
@app.route('/api/points/me', methods=['GET'])
@jwt_required()
def my_points():
    """返回当前用户的月度积分汇总与事件详情。
       可选参数：year_month=YYYY-MM（只看该月事件），否则返回最近 12 个月汇总 + 最近 50 条事件。"""
    user_id = int(get_jwt_identity())
    ym = request.args.get('year_month')

    # 汇总
    q_sum = MonthlyPoints.query.filter_by(user_id=user_id)
    if ym:
        q_sum = q_sum.filter(MonthlyPoints.year_month == ym)
    sums = q_sum.order_by(MonthlyPoints.year_month.desc()).limit(12).all()

    monthly = [{"year_month": s.year_month, "total_points": s.total_points} for s in sums]

    # 事件
    q_evt = PointsEvent.query.filter_by(user_id=user_id)
    if ym:
        q_evt = q_evt.filter(PointsEvent.year_month == ym)
    events = q_evt.order_by(PointsEvent.created_at.desc()).limit(50).all()

    evt_list = []
    for e in events:
        evt_list.append({
            "id": e.id,
            "year_month": e.year_month,
            "leaderboard_id": e.leaderboard_id,
            "submission_id": e.submission_id,
            "points": e.points,
            "delta": e.delta,
            "multiplier": e.multiplier,
            "difficulty_factor": e.difficulty_factor,
            "score": e.score,
            "sota_score": e.sota_score,
            "created_at": e.created_at.isoformat()
        })

    return jsonify({"monthly": monthly, "events": evt_list})


# ======================================================================
# 模块 6: Excel 导出（单任务 / 本榜单）
# ======================================================================

def _load_env_snapshot(sub: Submission):
    env = {}
    grid_keys = set()
    try:
        if sub.algo_env_json:
            env = json.loads(sub.algo_env_json) or {}
    except Exception:
        env = {}
    try:
        if sub.algo_env_grid_keys:
            arr = json.loads(sub.algo_env_grid_keys) or []
            grid_keys = {str(k) for k in arr}
    except Exception:
        grid_keys = set()
    return env, grid_keys


def _auto_resize_columns(ws):
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.rows:
        for cell in row:
            if cell.value is None:
                continue
            s = str(cell.value)
            l = len(s)
            col = cell.column
            widths[col] = max(widths.get(col, 0), l)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 60)


@app.route("/api/leaderboard/<int:leaderboard_id>/question_analysis", methods=["POST"])
def api_leaderboard_question_analysis(leaderboard_id: int):
    """
    POST /api/leaderboard/<id>/question_analysis

    Body:
    {
      "submission_ids": [12, 15, 19],              # 必填（建议前端传）
      "include_answers": true,                     # 默认 true
      "include_prompts": false,                    # 默认 false（prompt 很大）
      "include_answer_groups": true,               # 默认 true
      "answer_trunc": 800,                         # 默认 800（0 表示不截断）
      "sort_by": "q_rate_asc"                      # q_rate_asc|q_rate_desc|missing_desc|disagreement_desc|qid_asc
    }
    """
    import re
    import json
    from flask import request, jsonify

    # ---------- helpers ----------
    def _safe_int_list(v) -> list[int]:
        if v is None:
            return []
        if isinstance(v, (int, str)):
            v = [v]
        if not isinstance(v, (list, tuple)):
            return []
        out = []
        for x in v:
            try:
                out.append(int(x))
            except Exception:
                continue
        # 去重保持顺序
        seen = set()
        uniq = []
        for x in out:
            if x not in seen:
                uniq.append(x)
                seen.add(x)
        return uniq

    def _qid_sort_key(qid_raw):
        s = str(qid_raw).strip()
        if s.isdigit():
            return (0, int(s))
        m = re.search(r"(\d+)", s)
        if m:
            return (0, int(m.group(1)))
        return (1, s)

    def _norm_answer(text: str) -> str:
        if not text:
            return ""
        s = str(text).strip().lower()
        s = re.sub(r"\s+", " ", s)
        # 保留中英文数字下划线与空格（去掉大部分标点）
        s = re.sub(r"[^\w\u4e00-\u9fff ]+", "", s)
        return s[:240]  # 用于分桶，不要太长

    def _trunc(s: str, n: int) -> str:
        if not s:
            return ""
        if n and n > 0 and len(s) > n:
            return s[:n] + "…"
        return s

    # ---------- parse body ----------
    payload = request.get_json(silent=True) or {}
    submission_ids = _safe_int_list(payload.get("submission_ids"))

    include_answers = bool(payload.get("include_answers", True))
    include_prompts = bool(payload.get("include_prompts", False))
    include_answer_groups = bool(payload.get("include_answer_groups", True))
    answer_trunc = int(payload.get("answer_trunc", 800) or 0)
    sort_by = str(payload.get("sort_by", "q_rate_asc") or "q_rate_asc").strip()

    board = Leaderboard.query.get(leaderboard_id)
    if not board:
        return jsonify({"ok": False, "error": f"leaderboard {leaderboard_id} not found"}), 404

    if not submission_ids:
        return jsonify({"ok": False, "error": "submission_ids is required"}), 400

    # ---------- load submissions (must belong to this leaderboard) ----------
    subs_rows = (
        db.session.query(
            Submission.id.label("submission_id"),
            Submission.submission_name.label("submission_name"),
            Submission.score.label("score"),
            Submission.submitted_at.label("submitted_at"),
            User.username.label("username"),
        )
        .join(User, User.id == Submission.user_id)
        .filter(
            Submission.leaderboard_id == leaderboard_id,
            Submission.id.in_(submission_ids),
        )
        .all()
    )

    found_ids = {r.submission_id for r in subs_rows}
    ignored_submission_ids = [sid for sid in submission_ids if sid not in found_ids]
    if not subs_rows:
        return jsonify(
            {"ok": False, "error": "no valid submissions found for this leaderboard", "ignored_submission_ids": ignored_submission_ids}
        ), 400

    # 保持前端选择顺序
    subs_map = {r.submission_id: r for r in subs_rows}
    ordered_subs = [subs_map[sid] for sid in submission_ids if sid in subs_map]
    denom = len(ordered_subs)

    # ---------- load eval details ----------
    # 取全量后在内存里按 (question_key, submission_id) 覆盖为“最后一条”（与你 Excel 逻辑一致：order asc，后写覆盖）
    details = (
        SubmissionEvalDetail.query
        .filter(SubmissionEvalDetail.submission_id.in_([s.submission_id for s in ordered_subs]))
        .order_by(SubmissionEvalDetail.id.asc())
        .all()
    )

    # question_key -> {question_id, question, gold, per_sub{sub_id: d}}
    qmap: dict[str, dict] = {}

    def _mk_question_key(d: SubmissionEvalDetail) -> str:
        # 优先 question_id；没有就用 question 文本前缀；再不行用 detail-id（无法跨算法对齐，但至少不崩）
        if getattr(d, "question_id", None):
            return str(d.question_id)
        qtxt = (getattr(d, "question", "") or "").strip()
        if qtxt:
            return "q:" + qtxt[:64]
        return f"detail-{d.id}"

    for d in details:
        qkey = _mk_question_key(d)
        entry = qmap.get(qkey)
        if not entry:
            entry = {
                "question_id": getattr(d, "question_id", None) or qkey,
                "question": getattr(d, "question", "") or "",
                "gold": getattr(d, "gold_answer", "") or "",
                "per_sub": {},
            }
            qmap[qkey] = entry

        # 尽量补全 question/gold
        if not entry.get("question") and getattr(d, "question", None):
            entry["question"] = d.question
        if not entry.get("gold") and getattr(d, "gold_answer", None):
            entry["gold"] = d.gold_answer

        # 覆盖为最新（order asc -> 后者覆盖）
        entry["per_sub"][int(d.submission_id)] = d

    # ---------- compute per-question stats ----------
    questions_out = []
    per_submission_summary = {s.submission_id: {"correct_cnt": 0, "wrong_cnt": 0, "missing_cnt": 0, "unknown_cnt": 0} for s in ordered_subs}

    for qkey in sorted(qmap.keys(), key=_qid_sort_key):
        entry = qmap[qkey]
        per_sub = entry["per_sub"]

        correct_cnt = 0
        wrong_cnt = 0
        missing_cnt = 0
        unknown_cnt = 0

        # 答案分桶（用于一致性）
        groups = {}  # norm_answer -> {"count": int, "submission_ids": [..]}
        present_cnt = 0

        by_submission = {}

        for s in ordered_subs:
            sid = int(s.submission_id)
            d = per_sub.get(sid)

            if not d:
                missing_cnt += 1
                per_submission_summary[sid]["missing_cnt"] += 1
                by_submission[str(sid)] = {"is_correct": None, "missing": True}
                continue

            present_cnt += 1

            is_corr = getattr(d, "is_correct", None)
            if is_corr is None:
                unknown_cnt += 1
                per_submission_summary[sid]["unknown_cnt"] += 1
            elif bool(is_corr):
                correct_cnt += 1
                per_submission_summary[sid]["correct_cnt"] += 1
            else:
                wrong_cnt += 1
                per_submission_summary[sid]["wrong_cnt"] += 1

            item = {"is_correct": (1 if is_corr is True else 0 if is_corr is False else None), "missing": False}

            if include_answers:
                ans = getattr(d, "pred_answer", "") or ""
                item["answer"] = _trunc(ans, answer_trunc)

                if include_answer_groups:
                    key = _norm_answer(ans) or "<empty>"
                    g = groups.get(key)
                    if not g:
                        g = {"key": key, "count": 0, "submission_ids": []}
                        groups[key] = g
                    g["count"] += 1
                    g["submission_ids"].append(sid)

            if include_prompts:
                prompt = getattr(d, "eval_prompt", "") or ""
                item["prompt"] = _trunc(prompt, 2000)  # prompt 再截一次，防爆

            by_submission[str(sid)] = item

        q_rate = (correct_cnt / denom) if denom > 0 else None

        # 一致性：最大答案桶占 present_cnt 的比例（不含 missing）
        agreement = None
        answer_groups = None
        if include_answer_groups and present_cnt > 0 and groups:
            max_bucket = max(g["count"] for g in groups.values())
            agreement = max_bucket / present_cnt
            # 只回传 Top 6 桶，避免太大
            answer_groups = sorted(groups.values(), key=lambda x: (-x["count"], str(x["key"])))[:6]

        questions_out.append({
            "question_key": qkey,
            "question_id": entry.get("question_id", qkey),
            "question": entry.get("question", "") or "",
            "gold_answer": entry.get("gold", "") or "",
            "stats": {
                "denom": denom,
                "present_cnt": present_cnt,
                "correct_cnt": correct_cnt,
                "wrong_cnt": wrong_cnt,
                "missing_cnt": missing_cnt,
                "unknown_cnt": unknown_cnt,
                "q_rate": q_rate,
                "agreement": agreement,                 # 0..1（越大越一致）
                "disagreement": (1 - agreement) if agreement is not None else None,
            },
            "by_submission": by_submission,
            "answer_groups": answer_groups,
        })

    # ---------- sorting ----------
    if sort_by == "q_rate_desc":
        questions_out.sort(key=lambda x: (x["stats"]["q_rate"] is None, -(x["stats"]["q_rate"] or 0.0)))
    elif sort_by == "missing_desc":
        questions_out.sort(key=lambda x: -int(x["stats"]["missing_cnt"] or 0))
    elif sort_by == "disagreement_desc":
        questions_out.sort(key=lambda x: (x["stats"]["disagreement"] is None, -(x["stats"]["disagreement"] or 0.0)))
    elif sort_by == "qid_asc":
        questions_out.sort(key=lambda x: _qid_sort_key(x["question_key"]))
    else:
        # 默认：最难优先（q_rate_asc）
        questions_out.sort(key=lambda x: (x["stats"]["q_rate"] is None, (x["stats"]["q_rate"] if x["stats"]["q_rate"] is not None else 1e9)))

    # ---------- per-submission summary ----------
    # 这里 denom_questions 用题数（qmap 的 key 数）做分母；missing/unknown 也算“未正确”
    qn = len(qmap)
    subs_out = []
    for s in ordered_subs:
        sid = int(s.submission_id)
        stt = per_submission_summary.get(sid, {})
        acc = (stt.get("correct_cnt", 0) / qn) if qn > 0 else None
        subs_out.append({
            "submission_id": sid,
            "submission_name": getattr(s, "submission_name", "") or "",
            "username": getattr(s, "username", "") or "",
            "score": float(s.score) if s.score is not None else None,
            "submitted_at": (s.submitted_at.isoformat() if getattr(s, "submitted_at", None) else None),
            "summary": {
                "questions_total": qn,
                "correct_cnt": stt.get("correct_cnt", 0),
                "wrong_cnt": stt.get("wrong_cnt", 0),
                "missing_cnt": stt.get("missing_cnt", 0),
                "unknown_cnt": stt.get("unknown_cnt", 0),
                "acc_over_questions": acc,
            }
        })

    return jsonify({
        "ok": True,
        "leaderboard_id": leaderboard_id,
        "ignored_submission_ids": ignored_submission_ids,
        "denom": denom,
        "submissions": subs_out,
        "questions": questions_out,
    })



from io import BytesIO
from io import BytesIO

def _generate_single_submission_excel(submission: Submission) -> BytesIO:
    import xlsxwriter
    import re
    import json
    from typing import Any

    # ---------- 通用工具函数 ----------

    def _to_str_list(val: Any) -> list[str]:
        """把 None / str / list[...] 统一变成 List[str]，保留原始内容，后面再做显示清洗。"""
        if val is None:
            return []
        if isinstance(val, list):
            items = val
        else:
            items = [val]
        out: list[str] = []
        for v in items:
            if v is None:
                continue
            s = str(v)
            # 统一换行符，先保留，用于句子拆分
            s = s.replace("\r\n", "\n").replace("\r", "\n")
            if s.strip():
                out.append(s)
        return out

    # 展示/匹配用：压扁内容，去掉多余换行，避免“一格几百行”
    def _normalize_for_display(s: str) -> str:
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        # 把所有空白（包括换行）压成一个空格
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    # 去标点的匹配串构造（clean 文本 + clean 索引到原始索引映射）
    punct_pattern = re.compile(r"[，,。\.！？!?：:；;（）()\[\]【】“”\"‘’'·、…\-——_<>《》]")

    def _build_clean_index(s: str) -> tuple[str, list[int]]:
        clean_chars: list[str] = []
        idx_map: list[int] = []
        for i, ch in enumerate(s):
            if ch.isspace():
                continue
            if punct_pattern.search(ch):
                continue
            clean_chars.append(ch)
            idx_map.append(i)
        return "".join(clean_chars), idx_map

    def _extract_long_sentences_from_chunk(chunk: str, min_len: int = 6) -> list[str]:
        """
        按句号/问号/感叹号/分号/冒号/换行切分为“句子”，只保留长度 >= min_len 的。
        """
        if not chunk:
            return []
        chunk = chunk.replace("\r\n", "\n").replace("\r", "\n")
        parts = re.split(r"[。\.！？!?：:；;\n]+", chunk)
        out: list[str] = []
        for p in parts:
            p = p.strip()
            if len(p) >= min_len:
                out.append(p)
        return out

    def _merge_spans(spans: list[tuple[int, int]]) -> list[tuple[int, int]]:
        """合并重叠/相邻区间，得到不重叠的 [start, end) 列表。"""
        if not spans:
            return []
        spans_sorted = sorted(spans)
        merged: list[list[int]] = [[spans_sorted[0][0], spans_sorted[0][1]]]
        for s, e in spans_sorted[1:]:
            last_s, last_e = merged[-1]
            if s <= last_e:
                if e > last_e:
                    merged[-1][1] = e
            else:
                merged.append([s, e])
        return [(s, e) for s, e in merged]

    def _build_spans_with_clean(text_raw: str, sentences: list[str]) -> list[tuple[int, int]]:
        """
        给定原始展示文本 text_raw 和若干 sentences，使用「清洗-匹配-回映射」得到
        text_raw 里需要高亮的若干区间 spans。
        """
        clean_text, idx_map = _build_clean_index(text_raw)
        spans: list[tuple[int, int]] = []
        if not clean_text:
            return []
        for sen in sentences:
            clean_sen, _ = _build_clean_index(sen)
            if not clean_sen:
                continue
            start = 0
            while True:
                pos = clean_text.find(clean_sen, start)
                if pos == -1:
                    break
                raw_start = idx_map[pos]
                raw_end = idx_map[pos + len(clean_sen) - 1] + 1
                spans.append((raw_start, raw_end))
                start = pos + len(clean_sen)
        return _merge_spans(spans)

    def _update_width(widths: dict[int, int], col: int, value: Any):
        """简易列宽统计：按字符串长度估一个宽度。"""
        if value is None:
            return
        s = str(value)
        l = len(s)
        if l > widths.get(col, 0):
            widths[col] = l

    def _write_rich(ws, row: int, col: int, text: str,
                    spans: list[tuple[int, int]],
                    widths: dict[int, int],
                    hit_fmt,
                    cell_fmt=None):
        """
        在单元格内对 spans 指定的区间做红色富文本，高亮命中的部分。
        cell_fmt 用于设置 text_wrap / 对齐等。
        """
        if not text:
            ws.write(row, col, "", cell_fmt)
            _update_width(widths, col, "")
            return
        if not spans:
            ws.write(row, col, text, cell_fmt)
            _update_width(widths, col, text)
            return

        pieces = []
        last = 0
        for start, end in spans:
            if start > last:
                pieces.append(text[last:start])
            pieces.append((hit_fmt, text[start:end]))
            last = end
        if last < len(text):
            pieces.append(text[last:])

        args = []
        for p in pieces:
            if isinstance(p, tuple):
                fmt, s = p
                if s:
                    args.append(fmt)
                    args.append(s)
            else:
                s = p
                if s:
                    args.append(s)

        if len([x for x in args if isinstance(x, str)]) < 2:
            ws.write(row, col, text, cell_fmt)
        else:
            # 最后一个参数是 cell 的整体格式（例如 text_wrap）
            if cell_fmt is not None:
                ws.write_rich_string(row, col, *args, cell_fmt)
            else:
                ws.write_rich_string(row, col, *args)

        _update_width(widths, col, text)

    # ---------- 创建工作簿 ----------

    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})

    header_fmt = wb.add_format({"bold": True})
    hit_fmt = wb.add_format({"font_color": "red"})
    # Retrieved / Golden 专用：自动换行 + 顶端对齐
    wrap_cell_fmt = wb.add_format({"text_wrap": True, "valign": "top"})

    # ---------- Sheet 1: EvalResult（块内文字标红 + 自适应宽高） ----------

    ws_eval = wb.add_worksheet("EvalResult")
    widths_eval: dict[int, int] = {}

    details = (
        SubmissionEvalDetail.query
        .filter_by(submission_id=submission.id)
        .order_by(SubmissionEvalDetail.id.asc())
        .all()
    )

    # 收集 extra_json 的 key 作为 extra.* 列
    extra_keys: list[str] = []
    extra_key_set: set[str] = set()
    for d in details:
        if d.extra_json:
            try:
                obj = json.loads(d.extra_json)
                if isinstance(obj, dict):
                    for k in obj.keys():
                        if k not in extra_key_set:
                            extra_key_set.add(k)
                            extra_keys.append(k)
            except Exception:
                continue

    ragas_metric_names: list[str] = []
    ragas_metric_name_set: set[str] = set()
    for d in details:
        if d.extra_json:
            try:
                obj = json.loads(d.extra_json)
                rs = obj.get("ragas_scores") if isinstance(obj, dict) else None
                if isinstance(rs, dict):
                    for m in rs.keys():
                        if m not in ragas_metric_name_set:
                            ragas_metric_name_set.add(m)
                            ragas_metric_names.append(m)
            except Exception:
                continue

    headers = [
        "QID", "Question", "GoldAnswer", "AlgoAnswer",
        "IsCorrect", "LatencyMs", "UsedTokens", "retrievedCount",
        "RetrievedContents", "GoldenChunks", "EvalPrompt",
    ]
    headers.extend([f"extra.{k}" for k in extra_keys])
    headers.extend([f"CountedIn_{m}" for m in ragas_metric_names])

    # 表头
    for col, h in enumerate(headers):
        ws_eval.write(0, col, h, header_fmt)
        _update_width(widths_eval, col, h)

    try:
        retrieved_col_idx = headers.index("RetrievedContents")
    except ValueError:
        retrieved_col_idx = None
    try:
        golden_col_idx = headers.index("GoldenChunks")
    except ValueError:
        golden_col_idx = None

    row_idx = 1
    for d in details:
        # 解析 extra
        extra_obj: dict[str, Any] = {}
        if d.extra_json:
            try:
                extra_obj = json.loads(d.extra_json)
            except Exception:
                extra_obj = {}

        retrieved_count = d.retrieved if d.retrieved is not None else ""

        # 原始 list（用于句子拆分 & 语义）
        raw_ctx_list = _to_str_list(extra_obj.get("contexts"))
        raw_golden_list = _to_str_list(
            extra_obj.get("goldenchunks")
            or extra_obj.get("golden_chunks")
            or extra_obj.get("golden_contexts")
        )

        # 展示/匹配用：压扁换行，每块 normalize，然后用 " ||| " 拼接，防止行高爆炸
        disp_ctx_list = [_normalize_for_display(x) for x in raw_ctx_list]
        disp_golden_list = [_normalize_for_display(x) for x in raw_golden_list]

        retrieved_text = " ||| ".join(disp_ctx_list) if disp_ctx_list else ""
        golden_text = " ||| ".join(disp_golden_list) if disp_golden_list else ""

        # 从黄金块里抽出“句子”（长度>=6）——用原始块内容拆句
        golden_sents: list[str] = []
        for ch in raw_golden_list:
            golden_sents.extend(_extract_long_sentences_from_chunk(ch, min_len=6))

        # 先在 retrieved_text 里找命中的黄金句，收集命中的黄金句
        retrieved_spans: list[tuple[int, int]] = []
        hit_sents: list[str] = []

        if retrieved_text and golden_sents:
            clean_ret, idx_map_ret = _build_clean_index(retrieved_text)
            if clean_ret:
                for sen in golden_sents:
                    clean_sen, _ = _build_clean_index(sen)
                    if not clean_sen:
                        continue
                    start = 0
                    found = False
                    while True:
                        pos = clean_ret.find(clean_sen, start)
                        if pos == -1:
                            break
                        found = True
                        raw_start = idx_map_ret[pos]
                        raw_end = idx_map_ret[pos + len(clean_sen) - 1] + 1
                        retrieved_spans.append((raw_start, raw_end))
                        start = pos + len(clean_sen)
                    if found:
                        hit_sents.append(sen)

        retrieved_spans = _merge_spans(retrieved_spans)

        # 在 golden_text 自己里，对“命中的黄金句”做标红
        golden_spans: list[tuple[int, int]] = []
        if golden_text and hit_sents:
            golden_spans = _build_spans_with_clean(golden_text, hit_sents)

        # 基本列数据
        base_row_vals = [
            d.question_id or "",
            d.question or "",
            d.gold_answer or "",
            d.pred_answer or "",
            (1 if d.is_correct else 0) if d.is_correct is not None else "",
            d.latency_ms if d.latency_ms is not None else "",
            d.used_tokens if d.used_tokens is not None else "",
            retrieved_count,
            retrieved_text,
            golden_text,
            d.eval_prompt or "",
        ]

        # extra.* 列
        extra_vals: list[Any] = []
        for k in extra_keys:
            v = extra_obj.get(k, "")
            if isinstance(v, (list, dict, set, tuple)):
                try:
                    v = json.dumps(v, ensure_ascii=False)
                except Exception:
                    v = str(v)
            extra_vals.append(v)

        full_vals = base_row_vals + extra_vals
        ragas_scores_obj = extra_obj.get("ragas_scores") or {}
        counted_in_vals: list[Any] = []
        for m in ragas_metric_names:
            val = ragas_scores_obj.get(m)
            # val 不为 None 且是 float → main.py 评测成功 → 参与了分母
            counted_in_vals.append(1 if (val is not None and isinstance(val, (int, float))) else 0)

        full_vals = base_row_vals + extra_vals + counted_in_vals

        # 写行；RetrievedContents / GoldenChunks 用富文本 + 自动换行；其他普通写
        for col, val in enumerate(full_vals):
            if retrieved_col_idx is not None and col == retrieved_col_idx:
                _write_rich(
                    ws_eval,
                    row_idx,
                    col,
                    retrieved_text,
                    retrieved_spans,
                    widths_eval,
                    hit_fmt,
                    wrap_cell_fmt,
                )
            elif golden_col_idx is not None and col == golden_col_idx:
                _write_rich(
                    ws_eval,
                    row_idx,
                    col,
                    golden_text,
                    golden_spans,
                    widths_eval,
                    hit_fmt,
                    wrap_cell_fmt,
                )
            else:
                ws_eval.write(row_idx, col, val)
                _update_width(widths_eval, col, val)

        # —— 行高自适应：按当前行的 retrieved/golden 长度估算 —— #
        max_text_len = max(len(retrieved_text), len(golden_text))
        if max_text_len > 0:
            # 假设一行大约放 80 个字符，估算行数，最多 10 行
            approx_lines = min(max_text_len // 80 + 1, 10)
            base_height = 15  # 1 行约 15pt
            ws_eval.set_row(row_idx, base_height * approx_lines)

        row_idx += 1

    # 调整列宽：
    # - 普通列：按内容长度估算，最大 40
    # - RetrievedContents / GoldenChunks：放宽到最大 80，更“扁平”一点
    for col, w in widths_eval.items():
        base_w = w + 2
        if retrieved_col_idx is not None and col == retrieved_col_idx:
            ws_eval.set_column(col, col, min(base_w, 80))
        elif golden_col_idx is not None and col == golden_col_idx:
            ws_eval.set_column(col, col, min(base_w, 80))
        else:
            ws_eval.set_column(col, col, min(base_w, 40))

    # ---------- Sheet 2: Env ----------

    ws_env = wb.add_worksheet("Env")
    widths_env: dict[int, int] = {}

    env_headers = ["Key", "Value", "Type"]
    for col, h in enumerate(env_headers):
        ws_env.write(0, col, h, header_fmt)
        _update_width(widths_env, col, h)

    env, grid_keys = _load_env_snapshot(submission)
    row_idx = 1
    for k in sorted(env.keys()):
        v = env.get(k, "")
        typ = "grid" if k in grid_keys else "fixed"
        row_vals = [k, str(v), typ]
        for col, val in enumerate(row_vals):
            ws_env.write(row_idx, col, val)
            _update_width(widths_env, col, val)
        row_idx += 1

    for col, w in widths_env.items():
        ws_env.set_column(col, col, min(w + 2, 60))

    # ---------- Sheet 3: Leaderboard ----------

    ws_lb = wb.add_worksheet("Leaderboard")
    widths_lb: dict[int, int] = {}

    lb_headers = [
        "Rank", "SubmissionID", "User", "SubmissionName",
        "JobName", "Image", "Score", "SubmittedAt",
    ]
    for col, h in enumerate(lb_headers):
        ws_lb.write(0, col, h, header_fmt)
        _update_width(widths_lb, col, h)

    q = (
        db.session.query(
            Submission.id.label("submission_id"),
            Submission.submission_name,
            Submission.k8s_job_name,
            Submission.algorithm_image_url,
            Submission.score,
            Submission.submitted_at,
            User.username,
        )
        .join(User, User.id == Submission.user_id)
        .filter(
            Submission.leaderboard_id == submission.leaderboard_id,
            Submission.status == "Succeeded",
            Submission.score.isnot(None),
            Submission.score > 0.0,
        )
        .order_by(db.desc(Submission.score), db.asc(Submission.submitted_at))
    )
    rows = q.all()
    top_submission_id = rows[0].submission_id if rows else None

    top_fmt = wb.add_format({"bg_color": "#FDE68A"})
    mine_fmt = wb.add_format({"bg_color": "#BBF7D0"})

    row_idx = 1
    for idx, r in enumerate(rows, start=1):
        base_vals = [
            idx,
            r.submission_id,
            r.username,
            r.submission_name,
            r.k8s_job_name,
            r.algorithm_image_url,
            r.score,
            r.submitted_at.isoformat() if r.submitted_at else "",
        ]

        row_fmt = None
        if r.submission_id == top_submission_id:
            row_fmt = top_fmt
        if r.submission_id == submission.id:
            row_fmt = mine_fmt

        for col, val in enumerate(base_vals):
            ws_lb.write(row_idx, col, val, row_fmt)
            _update_width(widths_lb, col, val)
        row_idx += 1

    for col, w in widths_lb.items():
        ws_lb.set_column(col, col, min(w + 2, 40))

    # ---------- Sheet 4: Metrics ----------

    ws_metrics = wb.add_worksheet("Metrics")
    widths_metrics: dict[int, int] = {}

    metrics_obj = {}
    metrics_raw = None
    for attr in ("metrics_json", "metrics", "metrics_raw"):
        if hasattr(submission, attr):
            metrics_raw = getattr(submission, attr)
            break

    if metrics_raw:
        try:
            if isinstance(metrics_raw, str):
                metrics_obj = json.loads(metrics_raw)
            else:
                metrics_obj = metrics_raw or {}
        except Exception:
            metrics_obj = {}

    if metrics_obj:
        flat: dict[str, Any] = {}

        for k, v in metrics_obj.items():
            if k in ("score_detail", "ragas_summary", "ragas_config", "ragas"):
                continue
            if isinstance(v, (list, dict, tuple, set)):
                try:
                    flat[k] = json.dumps(v, ensure_ascii=False)
                except Exception:
                    flat[k] = str(v)
            else:
                flat[k] = v

        score_detail = metrics_obj.get("score_detail") or {}
        if isinstance(score_detail, dict):
            for k, v in score_detail.items():
                flat[f"score_detail.{k}"] = v

        ragas = metrics_obj.get("ragas") or {}
        if isinstance(ragas, dict):
            metrics_used = ragas.get("metrics_used")
            if metrics_used:
                if isinstance(metrics_used, (list, tuple, set)):
                    val = ", ".join(str(x) for x in metrics_used)
                else:
                    val = str(metrics_used)
                flat["ragas.metrics_used"] = val

            avg_scores = ragas.get("average_scores") or {}
            if isinstance(avg_scores, dict):
                for k, v in avg_scores.items():
                    flat[f"ragas.average_scores.{k}"] = v

        headers_metrics = list(sorted(flat.keys()))
        for col, h in enumerate(headers_metrics):
            ws_metrics.write(0, col, h, header_fmt)
            _update_width(widths_metrics, col, h)

        row_vals = []
        for k in headers_metrics:
            v = flat.get(k, "")
            if isinstance(v, (list, dict, tuple, set)):
                try:
                    v = json.dumps(v, ensure_ascii=False)
                except Exception:
                    v = str(v)
            row_vals.append(v)

        for col, val in enumerate(row_vals):
            ws_metrics.write(1, col, val)
            _update_width(widths_metrics, col, val)

        for col, w in widths_metrics.items():
            ws_metrics.set_column(col, col, min(w + 2, 60))
    else:
        ws_metrics.write(0, 0, "info", header_fmt)
        ws_metrics.write(1, 0, "No metrics JSON found for this submission.")

    # ---------- 收尾 ----------

    wb.close()
    buf.seek(0)
    return buf

def _generate_leaderboard_excel(board: Leaderboard, submissions: list[Submission]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.cell.cell import MergedCell
    import re

    wb = Workbook()

    # Sheet 1: 评测结果（多 algo，按 question 对齐）
    ws_eval = wb.active
    ws_eval.title = "EvalResult"

    # 过滤 score<=0，仅对有效提交做汇总
    valid_subs = [s for s in submissions if s.score is not None and s.score > 0.0]
    if not valid_subs:
        valid_subs = submissions

    sub_ids = [s.id for s in valid_subs]

    details = (
        SubmissionEvalDetail.query
        .filter(SubmissionEvalDetail.submission_id.in_(sub_ids))
        .order_by(SubmissionEvalDetail.id.asc())
        .all()
    )

    # question_id -> {question, gold, per_submission{sub_id: detail}}
    qmap: dict[str, dict] = {}
    for d in details:
        key = d.question_id or f"row-{len(qmap) + 1}"
        entry = qmap.get(key)
        if not entry:
            entry = {"question": d.question, "gold": d.gold_answer, "per_sub": {}}
            qmap[key] = entry
        # 如果后续 detail 带来更完整的 question/gold，可以覆盖
        if not entry.get("question") and d.question:
            entry["question"] = d.question
        if not entry.get("gold") and d.gold_answer:
            entry["gold"] = d.gold_answer
        entry["per_sub"][d.submission_id] = d

    # --- 新增：QID 数值排序辅助函数 ---
    def _qid_sort_key(qid_raw) -> tuple[int, int | str]:
        """
        让 QID 按“真实数字”排序：
        - 纯数字：按 int(qid) 排（1,2,10）
        - 含数字前缀：尽量提取前缀数字，例如 "Q12" -> 12, "row-3" -> 3
        - 完全非数字：放在后面，按字符串排序
        返回 (group, key)：group 保证“有数字”的在前，没有数字的在后。
        """
        s = str(qid_raw).strip()

        # 1) 纯数字
        if s.isdigit():
            return (0, int(s))

        # 2) 尝试抓前缀数字（比如 "Q12" / "row-3" / "id_007"）
        m = re.search(r"(\d+)", s)
        if m:
            return (0, int(m.group(1)))

        # 3) 完全没有数字的，排在后面，按字符串
        return (1, s)

    # --- 新增：按题统计“横向正确率”（该题在所有算法上的正确率）并按红->蓝渐变给整行上色 ---
    def _lerp(a: int, b: int, t: float) -> int:
        return int(round(a + (b - a) * t))

    def _mix_hex(c1: str, c2: str, t: float) -> str:
        """
        c1/c2: 'RRGGBB'（不带#）
        t: 0..1
        """
        t = 0.0 if t < 0 else (1.0 if t > 1 else t)
        r1, g1, b1 = int(c1[0:2], 16), int(c1[2:4], 16), int(c1[4:6], 16)
        r2, g2, b2 = int(c2[0:2], 16), int(c2[2:4], 16), int(c2[4:6], 16)
        return f"{_lerp(r1, r2, t):02X}{_lerp(g1, g2, t):02X}{_lerp(b1, b2, t):02X}"

    # 低正确率：浅红；高正确率：浅蓝（整行淡色，不刺眼）
    LOW_COLOR = "FEE2E2"   # light red
    HIGH_COLOR = "DBEAFE"  # light blue

    # 表头：QID / Question / Gold / QCorrectRate + 每个 submission 三列（Answer/IsCorrect/Prompt）
    headers = ["QID", "Question", "GoldAnswer", "QCorrectRate"]
    for sub in valid_subs:
        label = f"{sub.submission_name}"
        headers.append(f"{label} - Answer")
        headers.append(f"{label} - Correct")
        headers.append(f"{label} - Prompt")
    ws_eval.append(headers)
    for cell in ws_eval[1]:
        cell.font = Font(bold=True)

    # 先按“数值排序”的 QID 顺序写每题数据
    for qid in sorted(qmap.keys(), key=_qid_sort_key):
        entry = qmap[qid]
        per_sub = entry["per_sub"]

        # 统计该题“横向正确率”：在所有 valid_subs 上（缺失/None 视作不正确）
        denom = len(valid_subs) if len(valid_subs) > 0 else 0
        correct_cnt = 0
        for sub in valid_subs:
            d = per_sub.get(sub.id)
            if d and d.is_correct is not None and bool(d.is_correct):
                correct_cnt += 1
        q_rate = (correct_cnt / denom) if denom > 0 else None  # 0..1

        # QID 列：保留原样（字符串），只是排序是按数值排好的
        row = [qid, entry.get("question", ""), entry.get("gold", "")]
        # 正确率显示单独数字（百分比）
        row.append(q_rate if q_rate is not None else "")

        for sub in valid_subs:
            d = per_sub.get(sub.id)
            if d:
                ans = d.pred_answer or ""
                corr = ""
                if d.is_correct is not None:
                    corr = 1 if d.is_correct else 0
                prompt = d.eval_prompt or ""
                row.extend([ans, corr, prompt])
            else:
                row.extend(["", "", ""])
        ws_eval.append(row)

        # 给“该题这一整行”上色（按正确率红->蓝渐变）
        excel_row = ws_eval.max_row
        if q_rate is not None:
            fill_hex = _mix_hex(LOW_COLOR, HIGH_COLOR, float(q_rate))
            fill = PatternFill("solid", fgColor=fill_hex)
            for c in ws_eval[excel_row]:
                # 跳过合并单元格占位（防 openpyxl 写入异常）
                if isinstance(c, MergedCell):
                    continue
                c.fill = fill

        # 正确率列设置百分比格式（第4列）
        rate_cell = ws_eval.cell(row=excel_row, column=4)
        if q_rate is not None:
            rate_cell.number_format = "0.0%"

    _auto_resize_columns(ws_eval)

    # Sheet 2: Leaderboard（和前端 per-submission 类似，过滤 score<=0）
    ws_lb = wb.create_sheet(title="Leaderboard")
    ws_lb.append([
        "Rank", "SubmissionID", "User", "SubmissionName",
        "JobName", "Image", "Score", "SubmittedAt"
    ])
    for cell in ws_lb[1]:
        cell.font = Font(bold=True)

    q = (
        db.session.query(
            Submission.id.label('submission_id'),
            Submission.submission_name,
            Submission.k8s_job_name,
            Submission.algorithm_image_url,
            Submission.score,
            Submission.submitted_at,
            User.username
        )
        .join(User, User.id == Submission.user_id)
        .filter(
            Submission.leaderboard_id == board.id,
            Submission.status == 'Succeeded',
            Submission.score.isnot(None),
            Submission.score > 0.0
        )
        .order_by(db.desc(Submission.score), db.asc(Submission.submitted_at))
    )
    rows = q.all()
    top_submission_id = rows[0].submission_id if rows else None

    top_fill = PatternFill("solid", fgColor="FFFDE68A")

    for idx, r in enumerate(rows, start=1):
        ws_lb.append([
            idx,
            r.submission_id,
            r.username,
            r.submission_name,
            r.k8s_job_name,
            r.algorithm_image_url,
            r.score,
            r.submitted_at.isoformat() if r.submitted_at else ""
        ])
        if r.submission_id == top_submission_id:
            excel_row = idx + 1
            for c in ws_lb[excel_row]:
                c.fill = top_fill

    _auto_resize_columns(ws_lb)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route('/api/submission/<int:sub_id>/excel', methods=['GET'])
@jwt_required()
def download_submission_excel(sub_id):
    """
    单 algo 任务 Excel：
    - EvalResult: 本任务每道题的 QA + 正误 + prompt + extra.*
    - Env: 本任务注入的 ENV（含 grid/fixed 标记）
    - Leaderboard: 当前榜单 per-submission 排名（过滤 score<=0，top1+当前任务高亮）
    """
    sub = Submission.query.get_or_404(sub_id)

    if not is_admin_or_owner(sub.user_id):
        return jsonify({"msg": "Forbidden: You are not the owner or an admin"}), 403

    if sub.status != 'Succeeded':
        return jsonify({"msg": f"Submission status is '{sub.status}', Excel only for Succeeded ones."}), 400

    # ---- 文件名安全化：保留中文，替换非法字符，压缩空白，限长 ----
    def _safe_filename_part(s: str, max_len: int = 60) -> str:
        if not s:
            return ""
        s = str(s).strip()
        s = re.sub(r"\s+", " ", s)
        # Windows/HTTP header 不友好字符 + 控制字符
        s = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", s)
        s = s.strip(" ._")
        return s[:max_len] if s else ""

    # ✅ 榜单名（以榜单名开头）
    lb_name = _safe_filename_part(sub.leaderboard.name if sub.leaderboard else f"leaderboard_{sub.leaderboard_id}")
    sub_name = _safe_filename_part(sub.submission_name or "submission")

    # ✅ 时间戳（尾部）
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    # 如果你更想用服务器本地时间（可能是东八区），改成：datetime.now().strftime(...)

    # <LeaderboardName>_<SubmissionName>_<ID>_<TIMESTAMP>.xlsx
    filename = f"{lb_name}_{sub_name}_{sub.id}_{ts}.xlsx"

    buf = _generate_single_submission_excel(sub)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/api/leaderboard/<int:leaderboard_id>/excel', methods=['GET'])
@jwt_required()
def download_leaderboard_excel(leaderboard_id):
    """
    本榜单 Excel：
    - EvalResult: 每道题 × 各个算法的答案/正误/prompt（按 submission_name 展开列）
    - Leaderboard: 本榜单 per-submission 排名（过滤 score<=0）
    """
    board = Leaderboard.query.get_or_404(leaderboard_id)

    subs = (
        Submission.query
        .filter(
            Submission.leaderboard_id == leaderboard_id,
            Submission.status == 'Succeeded',
            Submission.score.isnot(None)
        )
        .order_by(Submission.score.desc(), Submission.submitted_at.asc())
        .all()
    )
    if not subs:
        return jsonify({"msg": "No succeeded submissions for this leaderboard."}), 400

    buf = _generate_leaderboard_excel(board, subs)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = board.name.replace("/", "_").replace("\\", "_")
    filename = f"{safe_name}_{ts}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ===== Agent integration (agno) =====
from agent.runner import init_agent_routes

_agent_runner = init_agent_routes(
    app, db,
    User=User,
    Leaderboard=Leaderboard,
    Submission=Submission,
    SubmissionEvalDetail=SubmissionEvalDetail,
    EnvPreset=EnvPreset,
    SubmissionLog=SubmissionLog,
    k8s_core_v1=k8s_core_v1,
    k8s_batch_v1=k8s_batch_v1,
    K8S_NAMESPACE=K8S_NAMESPACE,
    sync_submission_status=_sync_submission_status,
    persist_submission_logs=_persist_submission_logs,
    start_k8s_job=_start_k8s_job,
    database_url=app.config["SQLALCHEMY_DATABASE_URI"],
)


# ======================================================================
# 启动
# ======================================================================
if __name__ == '__main__':
    port = int(os.environ.get("FLASK_RUN_PORT", 8000))
    debug_mode = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    # 若全新部署请先运行： flask --app app.py init-db --create-defaults
    app.run(host='0.0.0.0', port=port, debug=debug_mode)

